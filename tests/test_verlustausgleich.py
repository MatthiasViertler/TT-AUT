"""Tests for Verlustausgleich year-over-year tracking (JSON snapshot save/load + Overview sheet)."""
import json
from decimal import Decimal
from pathlib import Path

import pytest

from core.models import TaxSummary
from generators.writer import _load_history, _save_summary_json, _fill_overview_sheet

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _make_summary(year: int, person: str = "Test",
                  dividends: float = 100.0, gains: float = 200.0,
                  losses: float = 50.0, net: float = 250.0,
                  kest: float = 68.75, wht: float = 15.0,
                  remaining: float = 53.75,
                  dom_gains: float = 0.0, dom_losses: float = 0.0,
                  fgn_gains: float = 0.0, fgn_losses: float = 0.0) -> TaxSummary:
    s = TaxSummary(tax_year=year, person_label=person)
    s.total_dividends_eur = Decimal(str(dividends))
    s.total_gains_eur     = Decimal(str(gains))
    s.total_losses_eur    = Decimal(str(losses))
    s.net_taxable_eur     = Decimal(str(net))
    s.kest_due_eur        = Decimal(str(kest))
    s.wht_creditable_eur  = Decimal(str(wht))
    s.kest_remaining_eur  = Decimal(str(remaining))
    s.kz_981 = Decimal(str(dom_gains))
    s.kz_891 = Decimal(str(dom_losses))
    s.kz_994 = Decimal(str(fgn_gains))
    s.kz_892 = Decimal(str(fgn_losses))
    return s


def test_save_summary_json_creates_file(tmp_path):
    s = _make_summary(2024)
    p = tmp_path / "Test_2024_summary.json"
    _save_summary_json(s, p)
    assert p.exists()


def test_save_summary_json_correct_values(tmp_path):
    s = _make_summary(2024, dividends=123.45, gains=678.90)
    p = tmp_path / "Test_2024_summary.json"
    _save_summary_json(s, p)

    data = json.loads(p.read_text())
    assert data["tax_year"] == 2024
    assert data["person_label"] == "Test"
    assert Decimal(data["total_dividends_eur"]) == Decimal("123.45")
    assert Decimal(data["total_gains_eur"]) == Decimal("678.90")


def test_load_history_returns_sorted_years(tmp_path):
    for year in (2023, 2021, 2022):
        s = _make_summary(year)
        _save_summary_json(s, tmp_path / f"Test_{year}_summary.json")

    history = _load_history("Test", tmp_path)
    assert [e["tax_year"] for e in history] == [2021, 2022, 2023]


def test_load_history_filters_by_person(tmp_path):
    _save_summary_json(_make_summary(2024, person="Alice"),
                       tmp_path / "Alice_2024_summary.json")
    _save_summary_json(_make_summary(2024, person="Bob"),
                       tmp_path / "Bob_2024_summary.json")

    alice_hist = _load_history("Alice", tmp_path)
    assert len(alice_hist) == 1
    assert alice_hist[0]["person_label"] == "Alice"

    bob_hist = _load_history("Bob", tmp_path)
    assert len(bob_hist) == 1
    assert bob_hist[0]["person_label"] == "Bob"


def test_load_history_empty_dir(tmp_path):
    assert _load_history("Nobody", tmp_path) == []


def test_load_history_missing_dir():
    assert _load_history("Nobody", Path("/nonexistent/path")) == []


def test_load_history_skips_corrupt_json(tmp_path):
    # Valid file
    _save_summary_json(_make_summary(2024), tmp_path / "Test_2024_summary.json")
    # Corrupt file — should be silently skipped
    (tmp_path / "Test_2023_summary.json").write_text("not valid json")

    history = _load_history("Test", tmp_path)
    assert len(history) == 1
    assert history[0]["tax_year"] == 2024


def test_roundtrip_decimal_precision(tmp_path):
    s = _make_summary(2025, dividends=3808.73, kest=1047.40, wht=431.87, remaining=615.53)
    p = tmp_path / "Test_2025_summary.json"
    _save_summary_json(s, p)

    data = json.loads(p.read_text())
    assert Decimal(data["total_dividends_eur"]) == Decimal("3808.73")
    assert Decimal(data["kest_due_eur"]) == Decimal("1047.40")
    assert Decimal(data["kest_remaining_eur"]) == Decimal("615.53")


def test_kz_fields_saved_in_json(tmp_path):
    s = _make_summary(2025, dom_gains=500.0, dom_losses=100.0,
                      fgn_gains=9292.0, fgn_losses=2628.0)
    p = tmp_path / "Test_2025_summary.json"
    _save_summary_json(s, p)

    data = json.loads(p.read_text())
    assert Decimal(data["kz_981"]) == Decimal("500.0")
    assert Decimal(data["kz_891"]) == Decimal("100.0")
    assert Decimal(data["kz_994"]) == Decimal("9292.0")
    assert Decimal(data["kz_892"]) == Decimal("2628.0")


# ── Overview sheet rendering ──────────────────────────────────────────────────

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def _build_history(tmp_path, entries):
    """Save summaries to JSON and load back as history list."""
    for s in entries:
        _save_summary_json(s, tmp_path / f"{s.person_label}_{s.tax_year}_summary.json")
    return _load_history(entries[0].person_label, tmp_path)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_overview_sheet_column_count(tmp_path):
    history = _build_history(tmp_path, [_make_summary(2025)])
    wb = Workbook()
    ws = wb.active
    _fill_overview_sheet(ws, history, current_year=2025)
    # 11 visible columns (including Div YoY %) + supplementary chart cols up to 27
    assert ws.max_column >= 11


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_overview_sheet_dom_fgn_split(tmp_path):
    s = _make_summary(2025, dividends=500.0,
                      dom_gains=200.0, dom_losses=50.0,
                      fgn_gains=800.0, fgn_losses=300.0,
                      net=1150.0, kest=316.25, wht=75.0, remaining=241.25)
    history = _build_history(tmp_path, [s])
    wb = Workbook()
    ws = wb.active
    _fill_overview_sheet(ws, history, current_year=2025)

    # Row 3 = first data row (row 1 = title, row 2 = headers)
    assert ws.cell(3, 3).value == pytest.approx(200.0)   # Dom Gains KZ 981
    assert ws.cell(3, 4).value == pytest.approx(-50.0)   # Dom Losses KZ 891 (negative)
    assert ws.cell(3, 5).value == pytest.approx(800.0)   # Fgn Gains KZ 994
    assert ws.cell(3, 6).value == pytest.approx(-300.0)  # Fgn Losses KZ 892 (negative)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_overview_sheet_totals_row(tmp_path):
    s1 = _make_summary(2024, dom_gains=100.0, dom_losses=20.0,
                       fgn_gains=400.0, fgn_losses=80.0, net=400.0, kest=110.0,
                       wht=20.0, remaining=90.0)
    s2 = _make_summary(2025, dom_gains=150.0, dom_losses=30.0,
                       fgn_gains=600.0, fgn_losses=120.0, net=600.0, kest=165.0,
                       wht=30.0, remaining=135.0)
    history = _build_history(tmp_path, [s1, s2])
    wb = Workbook()
    ws = wb.active
    _fill_overview_sheet(ws, history, current_year=2025)

    # Totals row = row 5 (title + header + 2 data rows + totals)
    totals_row = 5
    assert ws.cell(totals_row, 3).value == pytest.approx(250.0)   # dom gains total
    assert ws.cell(totals_row, 4).value == pytest.approx(-50.0)   # dom losses total (negative)
    assert ws.cell(totals_row, 5).value == pytest.approx(1000.0)  # fgn gains total
    assert ws.cell(totals_row, 6).value == pytest.approx(-200.0)  # fgn losses total (negative)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_overview_sheet_empty_history(tmp_path):
    wb = Workbook()
    ws = wb.active
    _fill_overview_sheet(ws, [], current_year=2025)
    # Should render title + headers without error, no data rows
    assert ws.max_row >= 2


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_overview_sheet_current_year_highlighted(tmp_path):
    s = _make_summary(2025)
    history = _build_history(tmp_path, [s])
    wb = Workbook()
    ws = wb.active
    _fill_overview_sheet(ws, history, current_year=2025)

    # Current year cell should have LIGHT_FILL (DEEAF1), not plain white
    fill = ws.cell(3, 1).fill
    assert fill.patternType == "solid"
    assert fill.fgColor.rgb.upper().endswith("DEEAF1")   # LIGHT_FILL
