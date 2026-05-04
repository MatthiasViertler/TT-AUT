"""
Output module.
Generates four artefacts per run:

1.  {person}_{year}_transactions.csv   — full normalized transaction log
2.  {person}_{year}_tax_summary.txt    — E1kv Kennziffern, ready to copy into FinanzOnline
3.  {person}_{year}_dashboard.xlsx     — Excel workbook with dashboard + detail tabs
4.  {person}_{year}_freedom.html       — interactive financial independence dashboard

openpyxl is used for Excel (pure Python, no COM/xlwings dependency).
"""

import csv
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.models import NormalizedTransaction, TaxSummary, TransactionType
from output.freedom import write_freedom_html
from output.wht_reclaim import write_wht_reclaim_report

log = logging.getLogger(__name__)
ZERO = Decimal("0")


def write_all(
    transactions: list[NormalizedTransaction],
    summary: TaxSummary,
    output_dir: Path,
    config: dict,
) -> None:
    slug = f"{summary.person_label}_{summary.tax_year}"
    opts = config.get("output", {})

    if opts.get("csv", True):
        p = output_dir / f"{slug}_transactions.csv"
        _write_csv(transactions, p)
        print(f"  [out]    {p}")

    if opts.get("tax_summary", True):
        p = output_dir / f"{slug}_tax_summary.txt"
        _write_tax_summary(summary, p)
        print(f"  [out]    {p}")

    if opts.get("excel", True):
        if OPENPYXL_AVAILABLE:
            p = output_dir / f"{slug}_dashboard.xlsx"
            _write_excel(transactions, summary, p, config)
            print(f"  [out]    {p}")
        else:
            log.warning("openpyxl not installed — skipping Excel output. "
                        "Run: pip install openpyxl")

    if opts.get("html", True):
        p = output_dir / f"{slug}_freedom.html"
        write_freedom_html(transactions, summary, p, config)
        print(f"  [out]    {p}")

    if opts.get("wht_reclaim", True) and config.get("at_residency_start_year"):
        p = output_dir / f"{slug}_wht_reclaim.txt"
        write_wht_reclaim_report(transactions, config, summary.tax_year,
                                  summary.person_label, p)
        if p.exists() and p.stat().st_size > 0:
            print(f"  [out]    {p}")


# ── CSV ───────────────────────────────────────────────────────────────────────

HEADERS = [
    "trade_date", "settle_date", "broker", "txn_type", "symbol", "isin",
    "description", "country_code", "domicile", "asset_class",
    "quantity", "price", "price_currency",
    "orig_currency", "orig_amount",
    "commission", "commission_currency",
    "wht_rate_actual", "wht_amount_orig",
    "fx_rate_to_eur", "eur_amount", "eur_commission", "eur_wht",
    "notes",
]


def _write_csv(txns: list[NormalizedTransaction], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for t in sorted(txns, key=lambda x: x.trade_date):
            w.writerow([
                t.trade_date, t.settle_date, t.broker, t.txn_type.value,
                t.symbol, t.isin or "", t.description,
                t.country_code or "", t.domicile.value, t.asset_class.value,
                _fmt(t.quantity), _fmt(t.price), t.price_currency or "",
                t.orig_currency, _fmt(t.orig_amount),
                _fmt(t.commission), t.commission_currency or "",
                _fmt(t.wht_rate_actual), _fmt(t.wht_amount_orig),
                _fmt(t.fx_rate_to_eur), _fmt(t.eur_amount),
                _fmt(t.eur_commission), _fmt(t.eur_wht),
                t.notes,
            ])


def _fmt(v: Optional[Decimal]) -> str:
    if v is None:
        return ""
    return str(v.normalize())


# ── Tax Summary TXT ───────────────────────────────────────────────────────────

def _write_tax_summary(s: TaxSummary, path: Path) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [
        "=" * 62,
        f"  KAPITALERTRAG — E1kv KENNZIFFERN",
        f"  Person : {s.person_label}",
        f"  Year   : {s.tax_year}",
        f"  Created: {now}",
        "=" * 62,
        "",
        "  ENTER THESE VALUES INTO FINANZONLINE / E1kv FORM:",
        "",
        f"  ── 1.3.1 Dividenden + Zinsen ─────────────────────────────",
        f"  KZ 862   Inländ. Dividendenerträge          EUR {s.kz_862:>12,.2f}",
        f"  KZ 863   Ausländ. Dividendenerträge          EUR {s.kz_863:>12,.2f}",
        f"  ── 1.3.2 Kursgewinne / Kursverluste ──────────────────────",
        f"  KZ 981   Inländ. Kursgewinne (27,5%)        EUR {s.kz_981:>12,.2f}",
        f"  KZ 994   Ausländ. Kursgewinne (27,5%)        EUR {s.kz_994:>12,.2f}",
        f"  KZ 864   Inländ. Kursgewinne (25%)          EUR {s.kz_864:>12,.2f}",
        f"  KZ 865   Ausländ. Kursgewinne (25%)          EUR {s.kz_865:>12,.2f}",
        f"  KZ 891   Inländ. Kursverluste                EUR {s.kz_891:>12,.2f}",
        f"  KZ 892   Ausländ. Kursverluste               EUR {s.kz_892:>12,.2f}",
        f"  ── 1.3.4 Investmentfonds ─────────────────────────────────",
        f"  KZ 897   Inländ. Fonds-Ausschüttungen       EUR {s.kz_897:>12,.2f}",
        f"  KZ 898   Ausländ. Fonds-Ausschüttungen       EUR {s.kz_898:>12,.2f}",
        f"  KZ 936   Ausschüttungsgleiche Ertr. (Inl.)  EUR {s.kz_936:>12,.2f}",
        f"  KZ 937   Ausschüttungsgleiche Ertr. (Ausl.)  EUR {s.kz_937:>12,.2f}",
        f"  ── Saldo aus Punkt 1.3 ───────────────────────────────────",
        f"  Saldo Inländisch                             EUR {s.saldo_inland:>12,.2f}",
        f"  Saldo Ausländisch                            EUR {s.saldo_ausland:>12,.2f}",
        f"  ── 1.4 / 1.6 Bereits bezahlte Steuer ────────────────────",
        f"  KZ 899   KESt inländ. WP im Ausland          EUR {s.kz_899:>12,.2f}",
        f"  KZ 984   QSt inländ. Einkünfte (27,5%)      EUR {s.kz_984:>12,.2f}",
        f"  KZ 998   QSt ausländ. Einkünfte (27,5%)      EUR {s.kz_998:>12,.2f}",
        "",
        "─" * 62,
        f"  Net taxable income                           EUR {s.net_taxable_eur:>12,.2f}",
        f"  KESt due @ 27.5%                             EUR {s.kest_due_eur:>12,.2f}",
        f"  WHT creditable (max 15%)                     EUR {s.wht_creditable_eur:>12,.2f}",
        f"  KESt remaining to pay                        EUR {s.kest_remaining_eur:>12,.2f}",
        "─" * 62,
        "",
        f"  Transactions processed : {s.transaction_count}",
        f"  Unmatched sells        : {s.unmatched_sells}",
        f"  Missing FX rates       : {s.missing_fx_count}",
        "",
    ]

    if s.warnings:
        lines.append(f"  WARNINGS ({len(s.warnings)}):")
        for w in s.warnings:
            lines.append(f"    • {w}")
        lines.append("")

    lines += [
        "  NOTE: KZ 937 (Ausschüttungsgleiche Erträge) requires OeKB",
        "  fund data and is NOT automatically calculated. Check your",
        "  broker's tax report or oekb.at for any ETF/fund holdings.",
        "",
        "  This output is informational. Verify with your tax consultant.",
        "=" * 62,
    ]

    path.write_text("\n".join(lines), encoding="utf-8")


# ── Excel Dashboard ───────────────────────────────────────────────────────────

def _write_excel(txns: list[NormalizedTransaction],
                 summary: TaxSummary, path: Path, config: dict | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()

    # ── Tab 1: E1kv Summary (for tax consultant) ──────────────────────────────
    ws = wb.active
    ws.title = "E1kv Summary"
    _fill_summary_sheet(ws, summary)

    # ── Tab 2: All Transactions ───────────────────────────────────────────────
    wt = wb.create_sheet("Transactions")
    _fill_transactions_sheet(wt, txns, summary.tax_year)

    # ── Tab 3: Dividends only ─────────────────────────────────────────────────
    wd = wb.create_sheet("Dividends")
    div_txns = [t for t in txns
                if t.txn_type == TransactionType.DIVIDEND
                and t.trade_date.year == summary.tax_year]
    _fill_transactions_sheet(wd, div_txns, summary.tax_year, title="Dividends")

    # ── Tab 4: Trades only ────────────────────────────────────────────────────
    wtr = wb.create_sheet("Trades")
    trade_txns = [t for t in txns
                  if t.txn_type in (TransactionType.BUY, TransactionType.SELL)
                  and t.trade_date.year == summary.tax_year]
    _fill_transactions_sheet(wtr, trade_txns, summary.tax_year, title="Trades")

    # ── Tab 5: Freedom (static snapshot at config assumptions) ───────────────
    wf = wb.create_sheet("Freedom")
    _fill_freedom_sheet(wf, div_txns, summary, config or {})

    # ── Tab 6: Nichtmeldefonds (only if positions exist) ──────────────────────
    if summary.nichtmeldefonds:
        wnmf = wb.create_sheet("Nichtmeldefonds")
        _fill_nichtmeldefonds_sheet(wnmf, summary)

    wb.save(path)


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL   = "1F3864"   # dark navy — main title
SECTION_FILL  = "FFC000"   # amber — E1kv section headers (1.3.1, 1.3.2 …)
SUBSECT_FILL  = "FFE090"   # light amber — sub-labels within a section
ACCENT_FILL   = "2E75B6"   # mid blue — summary headers
LIGHT_FILL    = "DEEAF1"   # pale blue
SALDO_FILL    = "D6E4F7"   # light blue-grey — Saldo row
WARN_FILL     = "FFF2CC"   # yellow
GREEN_FILL    = "E2EFDA"
RED_FILL      = "FCE4D6"
WHITE         = "FFFFFF"


def _hfill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")


def _fill_summary_sheet(ws, s: TaxSummary) -> None:
    # Column layout: A=section, B=KZ, C=description, D=Inländisch, E=Ausländisch
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = [1]  # mutable counter

    def _r():
        r = row[0]
        row[0] += 1
        return r

    def title(text):
        r = _r()
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = _font(bold=True, color=WHITE, size=12)
        c.fill = _hfill(HEADER_FILL)
        c.alignment = _center()
        ws.row_dimensions[r].height = 24

    def col_headers():
        r = _r()
        for col, txt in [("A", ""), ("B", "KZ"), ("C", ""), ("D", "Inländisch EUR"), ("E", "Ausländisch EUR")]:
            c = ws[f"{col}{r}"]
            c.value = txt
            c.font = _font(bold=True, color=WHITE, size=10)
            c.fill = _hfill(ACCENT_FILL)
            c.alignment = _center()
            c.border = _border()

    def section(code, text):
        r = _r()
        ws[f"A{r}"] = code
        ws[f"A{r}"].font = _font(bold=True, size=10)
        ws[f"A{r}"].fill = _hfill(SECTION_FILL)
        ws[f"A{r}"].alignment = _center()
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]
        c.value = text
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(SECTION_FILL)
        ws.row_dimensions[r].height = 18

    def kz_row(kz_in, kz_out, label, val_in, val_out, fill=None, warn=False):
        r = _r()
        ws[f"A{r}"].fill = _hfill(fill or WHITE)
        ws[f"B{r}"] = f"{kz_in}/{kz_out}" if kz_in and kz_out else (kz_in or kz_out or "")
        ws[f"B{r}"].font = _font(bold=True, size=10)
        ws[f"B{r}"].alignment = _center()
        ws[f"B{r}"].fill = _hfill(fill or WHITE)
        ws[f"C{r}"] = label
        ws[f"C{r}"].font = _font(size=10, color="C00000" if warn else "000000")
        ws[f"C{r}"].fill = _hfill(WARN_FILL if warn else (fill or WHITE))
        for col, val in [("D", val_in), ("E", val_out)]:
            c = ws[f"{col}{r}"]
            if val is not None:
                c.value = float(val)
                c.number_format = '#,##0.00'
            c.alignment = _right()
            c.fill = _hfill(WARN_FILL if warn else (fill or WHITE))
            if val is not None and float(val) < 0:
                c.font = Font(color="C00000", size=10)
        for col in "ABCDE":
            ws[f"{col}{r}"].border = _border()

    def saldo_row(val_in, val_out):
        r = _r()
        ws.merge_cells(f"A{r}:C{r}")
        c = ws[f"A{r}"]
        c.value = "Saldo aus Punkt 1.3"
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(SALDO_FILL)
        c.alignment = _center()
        for col, val in [("D", val_in), ("E", val_out)]:
            cell = ws[f"{col}{r}"]
            cell.value = float(val)
            cell.number_format = '#,##0.00'
            cell.alignment = _right()
            cell.font = _font(bold=True, size=10,
                              color="C00000" if float(val) < 0 else "000000")
            cell.fill = _hfill(SALDO_FILL)
            cell.border = _border()
        for col in "ABC":
            ws[f"{col}{r}"].border = _border()
        ws.row_dimensions[r].height = 18

    def summary_row(label, val_in, val_out, fill=None):
        r = _r()
        ws.merge_cells(f"A{r}:C{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(fill or LIGHT_FILL)
        c.alignment = _center()
        for col, val in [("D", val_in), ("E", val_out)]:
            cell = ws[f"{col}{r}"]
            if val is not None:
                cell.value = float(val)
                cell.number_format = '#,##0.00'
            cell.alignment = _right()
            cell.font = _font(bold=True, size=10)
            cell.fill = _hfill(fill or LIGHT_FILL)
            cell.border = _border()
        for col in "ABC":
            ws[f"{col}{r}"].border = _border()
        ws.row_dimensions[r].height = 18

    def blank():
        r = _r()
        ws.row_dimensions[r].height = 5

    # ── Title + column headers ────────────────────────────────────────────────
    title(f"E1kv — Kapitalvermögen  |  {s.person_label}  |  {s.tax_year}")
    col_headers()

    # ── 1.3.1 Dividenden + Zinsen ─────────────────────────────────────────────
    section("1.3.1", "Einkünfte aus der Überlassung von Kapital "
            "(§ 27 Abs. 2 — Dividenden, Zinsen aus Wertpapieren, 27,5%)")
    kz_row("862", "863", "Dividendenerträge + Zinsen", s.kz_862, s.kz_863)
    blank()

    # ── 1.3.2 Kursgewinne ─────────────────────────────────────────────────────
    section("1.3.2", "Einkünfte aus realisierten Wertsteigerungen von Kapitalvermögen (§ 27 Abs. 3)")
    kz_row("981", "994", "Überschüsse — besonderer Steuersatz 27,5%",
           s.kz_981, s.kz_994, fill=GREEN_FILL if (s.kz_981 + s.kz_994) > 0 else None)
    kz_row("864", "865", "Überschüsse — besonderer Steuersatz 25% (Wertpapiere vor 2011)",
           s.kz_864, s.kz_865)
    kz_row("891", "892", "Verluste",
           -s.kz_891 if s.kz_891 else None, -s.kz_892 if s.kz_892 else None,
           fill=RED_FILL if (s.kz_891 + s.kz_892) > 0 else None)
    blank()

    # ── 1.3.3 Derivate ────────────────────────────────────────────────────────
    section("1.3.3", "Einkünfte aus verbrieften Derivaten (§ 27 Abs. 4)")
    kz_row("982", "993", "Überschüsse — 27,5%", s.kz_982, s.kz_993)
    kz_row("893", "894", "Überschüsse — 25%",   s.kz_893, s.kz_894)
    kz_row("895", "896", "Verluste",
           -s.kz_895 if s.kz_895 else None, -s.kz_896 if s.kz_896 else None,
           fill=RED_FILL if (s.kz_895 + s.kz_896) > 0 else None)
    blank()

    # ── 1.3.4 Investmentfonds ─────────────────────────────────────────────────
    section("1.3.4", "Einkünfte aus Investmentfonds und Immobilieninvestmentfonds (§ 27 Abs. 3+4 InvFG)")
    kz_row("897", "898", "Ausschüttungen — 27,5%", s.kz_897, s.kz_898)
    kz_row("936", "937", "Ausschüttungsgleiche Erträge — 27,5%  ⚠ OeKB-Daten erforderlich",
           s.kz_936, s.kz_937, warn=(s.kz_937 == 0))
    blank()

    # ── 1.3.5 Kryptowährungen ─────────────────────────────────────────────────
    section("1.3.5", "Einkünfte aus Kryptowährungen (§ 27b)")
    kz_row("171",  "",    "Laufende Einkünfte (Mining, Staking)", s.kz_171, None)
    kz_row("173",  "",    "Überschüsse aus Wertsteigerungen",     s.kz_173, None)
    kz_row("175",  "",    "Verluste",
           -s.kz_175 if s.kz_175 else None, None,
           fill=RED_FILL if s.kz_175 > 0 else None)
    blank()

    # ── Saldo 1.3 ─────────────────────────────────────────────────────────────
    saldo_row(s.saldo_inland, s.saldo_ausland)
    blank()

    # ── 1.4 KESt bereits bezahlt ──────────────────────────────────────────────
    section("1.4", "Kapitalertragssteuer, soweit sie auf die inländischen Kapitaleinküfte entfällt")
    kz_row("899", "", "KESt für inländ. WP, die im Ausland gehalten werden",
           s.kz_899, None, fill=LIGHT_FILL)
    blank()

    # ── 1.5 Abgeltungssteuer ──────────────────────────────────────────────────
    section("1.5", "Abgeltungssteuer nach den Steuerabkommen mit Liechtenstein")
    kz_row("942", "", "Abgeltungssteuer Liechtenstein", s.kz_942, None)
    blank()

    # ── 1.6 Anrechenbare QSt 27.5% ───────────────────────────────────────────
    section("1.6", "Anzurechnende ausländ. (Quellen)Steuer auf Einkünfte — Steuersatz 27,5%")
    kz_row("984", "998", "Anzurechnende Quellensteuer",
           s.kz_984, s.kz_998, fill=GREEN_FILL if s.kz_998 > 0 else None)
    blank()

    # ── 1.7 Anrechenbare QSt 25% ─────────────────────────────────────────────
    section("1.7", "Anzurechnende ausländ. (Quellen)Steuer auf Einkünfte — Steuersatz 25%")
    kz_row("900", "901", "Anzurechnende Quellensteuer 25%", s.kz_900, s.kz_901)
    blank()

    # ── Calculation summary ───────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "BERECHNUNG"
    c.font = _font(bold=True, color=WHITE, size=11)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 20

    net = s.saldo_inland + s.saldo_ausland
    summary_row("Steuerpflichtiger Gesamtbetrag (Saldo 1.3 gesamt)", net, None, LIGHT_FILL)
    summary_row("KESt (27,5%)", s.kest_due_eur, None)
    summary_row("Anzurechnende Quellensteuer (KZ 998)", s.wht_creditable_eur, None, GREEN_FILL)
    summary_row("Verbleibende KESt zu bezahlen", s.kest_remaining_eur, None, WARN_FILL)
    blank()

    # ── Notes ─────────────────────────────────────────────────────────────────
    notes_row = row[0]
    notes = [
        "⚠  KZ 936/937 (Ausschüttungsgleiche Erträge) werden NICHT automatisch berechnet — OeKB-Daten erforderlich.",
        "⚠  Nichtmeldefonds (REITs, BDCs): Sonderbehandlung — wird in einem eigenen Abschnitt ausgewiesen (noch nicht implementiert).",
        "    Verluste (KZ 891/892) werden hier als negative Werte dargestellt; in FinanzOnline als Absolutbeträge eintragen.",
        "    Diese Ausgabe ist informativ. Bitte mit Steuerberater:in abstimmen.",
    ]
    for i, note in enumerate(notes):
        r = row[0] + i
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = note
        c.font = Font(color="666666", italic=True, size=9)
        ws.row_dimensions[r].height = 14


def _fill_transactions_sheet(ws, txns: list[NormalizedTransaction],
                               year: int, title: str = "Transactions") -> None:
    headers = [
        "Date", "Type", "Symbol", "ISIN", "Description",
        "Country", "Domicile", "Qty", "Currency",
        "Orig Amount", "Commission", "WHT (orig)",
        "FX Rate", "EUR Amount", "EUR Commission", "EUR WHT",
        "Notes",
    ]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _hfill(HEADER_FILL)
        cell.alignment = _center()
        cell.border = _border()

    txns_sorted = sorted(txns, key=lambda t: t.trade_date)
    for t in txns_sorted:
        row = [
            t.trade_date.isoformat(),
            t.txn_type.value,
            t.symbol,
            t.isin or "",
            t.description[:60],
            t.country_code or "",
            t.domicile.value,
            float(t.quantity) if t.quantity else "",
            t.orig_currency,
            float(t.orig_amount),
            float(t.commission) if t.commission else "",
            float(t.wht_amount_orig) if t.wht_amount_orig else "",
            float(t.fx_rate_to_eur) if t.fx_rate_to_eur else "",
            float(t.eur_amount) if t.eur_amount else "",
            float(t.eur_commission) if t.eur_commission else "",
            float(t.eur_wht) if t.eur_wht else "",
            t.notes,
        ]
        ws.append(row)

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row_idx, col_idx).value or ""))
             for row_idx in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Number format for EUR columns
    eur_cols = [10, 11, 12, 14, 15, 16]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in eur_cols:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00'


# ── Freedom tab ──────────────────────────────────────────────────────────────

_FD_DEFAULTS = {
    "portfolio_eur": 10000,
    "monthly_expenses_eur": 2000,
    "monthly_contribution_eur": 500,
    "yield_pct": 3.0,
    "growth_pct": 7.0,
}

GREEN_DARK  = "375623"
BLUE_LIGHT2 = "BDD7EE"
PROJ_FILL   = "E7F0FC"


def _fill_freedom_sheet(
    ws,
    div_txns: list[NormalizedTransaction],
    summary: TaxSummary,
    config: dict,
) -> None:
    fd = {**_FD_DEFAULTS, **config.get("freedom_dashboard", {})}

    portfolio     = float(fd["portfolio_eur"])
    monthly_exp   = float(fd["monthly_expenses_eur"])
    monthly_cont  = float(fd["monthly_contribution_eur"])
    yield_pct     = float(fd["yield_pct"]) / 100.0
    growth_pct    = float(fd["growth_pct"]) / 100.0

    annual_div    = float(summary.total_dividends_eur)
    monthly_div   = annual_div / 12.0
    freedom_pct   = (monthly_div / monthly_exp * 100.0) if monthly_exp else 0.0

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    row = [1]

    def _r():
        r = row[0]; row[0] += 1; return r

    def _cell(r, col, value=None, bold=False, color="000000", size=11,
              fill=None, fmt=None, align="left", border=False):
        c = ws.cell(r, col, value)
        c.font = _font(bold=bold, color=color, size=size)
        if fill:
            c.fill = _hfill(fill)
        if fmt:
            c.number_format = fmt
        c.alignment = Alignment(horizontal=align, vertical="center")
        if border:
            c.border = _border()
        return c

    # ── Title ─────────────────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = f"Financial Freedom  |  {summary.person_label}  |  {summary.tax_year}"
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _hfill(HEADER_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 26

    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  "
               f"Actual dividends {summary.tax_year}  |  "
               f"Projection: {fd['yield_pct']}% yield · {fd['growth_pct']}% growth · "
               f"€{fd['monthly_contribution_eur']:,.0f}/mo contribution")
    c.font = _font(color="555555", size=9)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 14

    row[0] += 1  # blank spacer

    # ── Key Metrics ───────────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "KEY METRICS"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    metrics = [
        ("Annual Dividends (actual)",      annual_div,   '#,##0.00 "EUR"', None),
        ("Monthly Dividends (actual)",      monthly_div,  '#,##0.00 "EUR"', None),
        ("Monthly Expenses (config)",       monthly_exp,  '#,##0.00 "EUR"', None),
        ("Monthly Surplus / Deficit",       monthly_div - monthly_exp,
         '#,##0.00 "EUR"', GREEN_FILL if monthly_div >= monthly_exp else RED_FILL),
        ("Freedom %",                       freedom_pct,  '0.0"%"',          None),
        ("Portfolio Value (config)",        portfolio,    '#,##0 "EUR"',     LIGHT_FILL),
    ]
    for label, val, fmt, fill in metrics:
        r = _r()
        _cell(r, 1, label, bold=True, size=10, fill=fill or WHITE, border=True)
        c = _cell(r, 2, val, bold=True, size=10, fill=fill or WHITE,
                  fmt=fmt, align="right", border=True)
        if label == "Freedom %" :
            pct = min(val, 100.0)
            bar_fill = GREEN_FILL if pct >= 100 else (SECTION_FILL if pct >= 50 else LIGHT_FILL)
            for col in range(3, 6):
                ws.cell(r, col).fill = _hfill(bar_fill if col <= 2 + int(pct / 33.4) + 1 else WHITE)
        ws.row_dimensions[r].height = 16

    row[0] += 1  # blank

    # ── Holdings breakdown ────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = f"DIVIDEND HOLDINGS  —  {summary.tax_year}"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    # header
    r = _r()
    for col, txt in [(1, "Symbol"), (2, "Annual EUR"), (3, "Monthly EUR"), (4, "% of Total"), (5, "Payments")]:
        _cell(r, col, txt, bold=True, color=WHITE, size=9, fill=ACCENT_FILL, align="center", border=True)
    ws.row_dimensions[r].height = 15

    # aggregate per symbol
    acc: dict[str, dict] = {}
    for t in div_txns:
        if t.symbol not in acc:
            acc[t.symbol] = {"annual": ZERO, "payments": 0}
        acc[t.symbol]["annual"] += t.eur_amount or ZERO
        acc[t.symbol]["payments"] += 1

    total_div_dec = sum(v["annual"] for v in acc.values())

    for sym, data in sorted(acc.items(), key=lambda x: -x[1]["annual"]):
        r = _r()
        ann = float(data["annual"])
        pct = float(data["annual"] / total_div_dec * 100) if total_div_dec else 0
        _cell(r, 1, sym, bold=True, size=9, fill=WHITE, border=True)
        _cell(r, 2, ann, size=9, fill=WHITE, fmt='#,##0.00', align="right", border=True)
        _cell(r, 3, ann / 12, size=9, fill=WHITE, fmt='#,##0.00', align="right", border=True)
        _cell(r, 4, pct, size=9, fill=WHITE, fmt='0.0"%"', align="right", border=True)
        _cell(r, 5, data["payments"], size=9, fill=WHITE, align="center", border=True)
        ws.row_dimensions[r].height = 14

    # totals
    r = _r()
    _cell(r, 1, "TOTAL", bold=True, size=9, fill=LIGHT_FILL, border=True)
    _cell(r, 2, float(total_div_dec), bold=True, size=9, fill=LIGHT_FILL,
          fmt='#,##0.00', align="right", border=True)
    _cell(r, 3, float(total_div_dec) / 12, bold=True, size=9, fill=LIGHT_FILL,
          fmt='#,##0.00', align="right", border=True)
    _cell(r, 4, 100.0, bold=True, size=9, fill=LIGHT_FILL, fmt='0.0"%"',
          align="right", border=True)
    ws.row_dimensions[r].height = 14

    row[0] += 1  # blank

    # ── 10-Year Projection ────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "10-YEAR PROJECTION  (static at config assumptions)"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    r = _r()
    for col, txt in [(1, "Year"), (2, "Portfolio EUR"), (3, "Annual Divs EUR"),
                     (4, "Monthly Divs EUR"), (5, "Freedom %")]:
        _cell(r, col, txt, bold=True, color=WHITE, size=9, fill=ACCENT_FILL, align="center", border=True)
    ws.row_dimensions[r].height = 15

    port = portfolio
    for i in range(11):
        yr = summary.tax_year + i
        ann_d = port * yield_pct if i > 0 else annual_div
        mo_d  = ann_d / 12.0
        fp    = mo_d / monthly_exp * 100.0 if monthly_exp else 0.0
        fill  = GREEN_FILL if fp >= 100 else (PROJ_FILL if i % 2 == 0 else WHITE)

        r = _r()
        _cell(r, 1, yr,    bold=(i == 0), size=9, fill=fill, align="center", border=True)
        _cell(r, 2, port,  size=9, fill=fill, fmt='#,##0', align="right", border=True)
        _cell(r, 3, ann_d, size=9, fill=fill, fmt='#,##0.00', align="right", border=True)
        _cell(r, 4, mo_d,  size=9, fill=fill, fmt='#,##0.00', align="right", border=True)
        _cell(r, 5, fp,    bold=(fp >= 100), size=9, fill=fill, fmt='0.0"%"',
              align="right", border=True)
        ws.row_dimensions[r].height = 14

        port = port * (1 + growth_pct) + monthly_cont * 12

    # note
    r = _r() + 1
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = (f"Year 0 = actual dividends. Years 1–10: portfolio × {fd['yield_pct']}% yield, "
               f"{fd['growth_pct']}% annual growth, +€{fd['monthly_contribution_eur']:,.0f}/mo contribution. "
               f"Update portfolio_eur in config.local.yaml for accurate projections.")
    c.font = _font(color="888888", size=8)
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes = "A3"


# ── Nichtmeldefonds tab ───────────────────────────────────────────────────────

def _fill_nichtmeldefonds_sheet(ws, summary) -> None:
    from openpyxl.styles import Font as OFont

    ws.column_dimensions["A"].width = 8    # Symbol
    ws.column_dimensions["B"].width = 6    # Type
    ws.column_dimensions["C"].width = 28   # Name
    ws.column_dimensions["D"].width = 5    # Ccy
    ws.column_dimensions["E"].width = 10   # Shares
    ws.column_dimensions["F"].width = 12   # Jan 1 price
    ws.column_dimensions["G"].width = 12   # Dec 31 price
    ws.column_dimensions["H"].width = 11   # Annual gain/sh
    ws.column_dimensions["I"].width = 10   # AE/sh (90%)
    ws.column_dimensions["J"].width = 10   # AE/sh (10%)
    ws.column_dimensions["K"].width = 10   # AE/sh used
    ws.column_dimensions["L"].width = 13   # AE total native
    ws.column_dimensions["M"].width = 11   # FX Dec31
    ws.column_dimensions["N"].width = 13   # AE total EUR
    ws.column_dimensions["O"].width = 13   # KeSt EUR
    ws.column_dimensions["P"].width = 13   # Cost basis +

    headers = [
        "Symbol", "Type", "Name", "Ccy", "Shares",
        "Jan 1 Price", "Dec 31 Price", "Gain/Share",
        "AE 90%/sh", "AE 10%/sh", "AE/share",
        f"AE ({summary.tax_year} native)", "FX Dec31",
        "AE (EUR)", "KeSt 27.5% (EUR)", "Cost Basis +"
    ]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _hfill(SECTION_FILL)
        cell.alignment = _center()
        cell.border = _border()

    # Data rows
    for r in summary.nichtmeldefonds:
        def fv(v):
            return float(v) if v else None

        row_data = [
            r.symbol,
            r.fund_type,
            r.name,
            r.currency,
            float(r.shares_held),
            fv(r.price_jan1),
            fv(r.price_dec31),
            fv(r.annual_gain_per_share),
            fv(r.ae_90pct_per_share),
            fv(r.ae_10pct_per_share),
            fv(r.ae_per_share),
            fv(r.ae_total_native),
            fv(r.fx_dec31),
            fv(r.ae_total_eur),
            fv(r.kest_due_eur),
            fv(r.cost_basis_adj_eur),
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Formatting
        ws.cell(row_idx, 1).font = _font(bold=True)       # symbol bold
        for col_idx in [6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00'
        ws.cell(row_idx, 13).number_format = '0.0000'     # FX rate
        for col_idx in [14, 15, 16]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00 "EUR"'
            ws.cell(row_idx, col_idx).font = _font(bold=(col_idx == 15))
        ws.cell(row_idx, 15).fill = _hfill(WARN_FILL)     # KeSt highlight

        if r.warning:
            ws.cell(row_idx, 1).font = OFont(color="C00000", bold=True)

        for col_idx in range(1, 17):
            ws.cell(row_idx, col_idx).border = _border()

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(total_row, 13).value = "TOTAL"
    ws.cell(total_row, 13).font = _font(bold=True)
    ws.cell(total_row, 14).value = float(summary.nichtmeldefonds_ae_eur)
    ws.cell(total_row, 14).number_format = '#,##0.00 "EUR"'
    ws.cell(total_row, 14).font = _font(bold=True)
    ws.cell(total_row, 14).fill = _hfill(LIGHT_FILL)
    ws.cell(total_row, 15).value = float(summary.nichtmeldefonds_kest_eur)
    ws.cell(total_row, 15).number_format = '#,##0.00 "EUR"'
    ws.cell(total_row, 15).font = _font(bold=True)
    ws.cell(total_row, 15).fill = _hfill(WARN_FILL)
    for col_idx in [14, 15]:
        ws.cell(total_row, col_idx).border = _border()

    # Notes
    note_row = total_row + 2
    notes = [
        f"KZ 937 (Ausschüttungsgleiche Erträge Ausland) = {float(summary.nichtmeldefonds_ae_eur):,.2f} EUR  →  KeSt = {float(summary.nichtmeldefonds_kest_eur):,.2f} EUR",
        "Cost Basis + = Erhöhung der steuerlichen Anschaffungskosten um die AE (verhindert Doppelbesteuerung beim Verkauf).",
        "AE/share = max(90% × Jahresgewinn/Anteil,  10% × Jahresendkurs/Anteil)   per § 186 Abs. 2 InvFG.",
        "Preise (dec31_prices in config) manuell eintragen: Jahresschlusskurs lt. IBKR-Portfolioaufstellung oder Yahoo Finance.",
    ]
    for i, note in enumerate(notes):
        ws.merge_cells(f"A{note_row + i}:P{note_row + i}")
        c = ws.cell(note_row + i, 1)
        c.value = note
        c.font = OFont(color="666666", italic=True, size=9)

    ws.freeze_panes = "A2"

