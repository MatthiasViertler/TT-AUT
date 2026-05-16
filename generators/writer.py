"""
Output module.
Generates five artefacts per run:

1.  {person}_{year}_transactions.csv   — full normalized transaction log
2.  {person}_{year}_tax_summary.txt    — E1kv Kennziffern, ready to copy into FinanzOnline
3.  {person}_{year}_dashboard.xlsx     — Excel workbook with dashboard + detail tabs
4.  {person}_{year}_freedom.html       — interactive financial independence dashboard
5.  {person}_{year}_summary.json       — machine-readable snapshot for year-over-year tracking

openpyxl is used for Excel (pure Python, no COM/xlwings dependency).
"""

import csv
import json
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.models import NormalizedTransaction, TaxSummary, TransactionType
from generators.anv_checklist import write_anv_checklist
from generators.freedom import write_freedom_html
from generators.tax_efficiency import write_tax_efficiency_report
from generators.wht_reclaim import write_wht_reclaim_report

log = logging.getLogger(__name__)
ZERO = Decimal("0")


def write_all(
    transactions: list[NormalizedTransaction],
    summary: TaxSummary,
    output_dir: Path,
    config: dict,
    fx=None,
) -> None:
    slug = f"{summary.person_label}_{summary.tax_year}"
    opts = config.get("output", {})

    # Save machine-readable snapshot first so Excel Overview tab can include current year
    p = output_dir / f"{slug}_summary.json"
    _save_summary_json(summary, p, transactions)
    print(f"  [out]    {p}")

    if opts.get("csv", True):
        p = output_dir / f"{slug}_transactions.csv"
        _write_csv(transactions, p)
        print(f"  [out]    {p}")

    if opts.get("tax_summary", True):
        p = output_dir / f"{slug}_tax_summary.txt"
        _write_tax_summary(summary, p)
        print(f"  [out]    {p}")

    if opts.get("excel", True):
        if OPENPYXL_AVAILABLE:
            history = _load_history(summary.person_label, output_dir)
            p = output_dir / f"{slug}_dashboard.xlsx"
            _write_excel(transactions, summary, p, config, history)
            print(f"  [out]    {p}")
        else:
            log.warning("openpyxl not installed — skipping Excel output. "
                        "Run: pip install openpyxl")

    if opts.get("html", True):
        p = output_dir / f"{slug}_freedom.html"
        write_freedom_html(transactions, summary, p, config)
        print(f"  [out]    {p}")

    if opts.get("wht_reclaim", True) and config.get("at_residency_start_year"):
        p = output_dir / f"{slug}_wht_reclaim.txt"
        write_wht_reclaim_report(transactions, config, summary.tax_year,
                                  summary.person_label, p)
        if p.exists() and p.stat().st_size > 0:
            print(f"  [out]    {p}")

    if opts.get("anv_checklist", True) and config.get("anv"):
        p = output_dir / f"{slug}_anv_checklist.txt"
        write_anv_checklist(config, summary.tax_year, summary.person_label, p)
        if p.exists() and p.stat().st_size > 0:
            print(f"  [out]    {p}")

    if opts.get("tax_efficiency", True) and fx is not None and summary.nichtmeldefonds:
        p = output_dir / f"{slug}_tax_efficiency.txt"
        write_tax_efficiency_report(
            config=config,
            tax_year=summary.tax_year,
            all_transactions=transactions,
            nmf_results=summary.nichtmeldefonds,
            portfolio_positions=summary.portfolio_positions,
            fx=fx,
            output_path=p,
        )
        print(f"  [out]    {p}")


# ── Summary JSON (Verlustausgleich history) ───────────────────────────────────

_SUMMARY_FIELDS = [
    "total_dividends_eur", "total_gains_eur", "total_losses_eur",
    "net_taxable_eur", "kest_due_eur", "wht_creditable_eur", "kest_remaining_eur",
    "kz_862", "kz_863", "kz_981", "kz_994", "kz_891", "kz_892",
]


def _save_summary_json(summary: TaxSummary, path: Path,
                       transactions: "list | None" = None) -> None:
    data: dict = {
        "tax_year": summary.tax_year,
        "person_label": summary.person_label,
    }
    for field in _SUMMARY_FIELDS:
        data[field] = str(getattr(summary, field))

    # Optional computed fields — serialized as float strings or null
    for opt_field in ("portfolio_eur_computed", "ibkr_cash_eur", "dividend_yield_computed"):
        val = getattr(summary, opt_field, None)
        data[opt_field] = str(val) if val is not None else None

    # Interest income (always present; ZERO if no IBKR interest section found)
    data["interest_eur"] = str(summary.interest_eur)

    # Monthly breakdown (for dashboard trend charts)
    if transactions is not None:
        monthly_divs: dict[str, str] = {}
        monthly_tx: dict[str, int] = {}
        for txn in transactions:
            if txn.trade_date.year != summary.tax_year:
                continue
            key = f"{txn.trade_date.month:02d}"
            monthly_tx[key] = monthly_tx.get(key, 0) + 1
            if txn.txn_type == TransactionType.DIVIDEND:
                prev = Decimal(monthly_divs.get(key, "0"))
                monthly_divs[key] = str(prev + txn.eur_amount)
        data["monthly_dividends"] = monthly_divs
        data["monthly_transaction_counts"] = monthly_tx
        data["transaction_count"] = sum(monthly_tx.values())

    # portfolio_positions: serialize for the multi-year Overview tab
    data["portfolio_positions"] = [
        {
            "symbol": p.symbol,
            "name": p.name,
            "fund_type": p.fund_type,
            "currency": p.currency,
            "qty": str(p.qty),
            "is_synthetic": p.is_synthetic,
            "eur_value": str(p.eur_value),
            "dividends_eur": str(p.dividends_eur),
            "yield_pct": p.yield_pct,
            "portfolio_pct": p.portfolio_pct,
        }
        for p in summary.portfolio_positions
    ]

    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def _load_history(person_label: str, output_dir: Path) -> list[dict]:
    """Return all saved year snapshots for this person, sorted by year ascending.

    Searches the parent output directory (users/{person}/output/) so that
    prior-year summary.json files from other year subdirectories are included.
    Falls back to searching output_dir itself for non-standard layouts.
    """
    # If output_dir is users/{person}/output/{year}/, walk up to output root
    search_root = output_dir
    if output_dir.name.isdigit():
        search_root = output_dir.parent

    if not search_root.exists():
        return []

    seen: set[int] = set()
    entries = []
    for p in search_root.rglob(f"{person_label}_*_summary.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if data.get("person_label") == person_label:
                yr = data.get("tax_year")
                if yr not in seen:
                    seen.add(yr)
                    entries.append(data)
        except Exception:
            pass
    return sorted(entries, key=lambda x: x.get("tax_year", 0))


# ── CSV ───────────────────────────────────────────────────────────────────────

HEADERS = [
    "trade_date", "settle_date", "broker", "txn_type", "symbol", "isin",
    "description", "country_code", "domicile", "asset_class",
    "quantity", "price", "price_currency",
    "orig_currency", "orig_amount",
    "commission", "commission_currency",
    "wht_rate_actual", "wht_amount_orig",
    "fx_rate_to_eur", "eur_amount", "eur_commission", "eur_wht",
    "notes",
]


def _write_csv(txns: list[NormalizedTransaction], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for t in sorted(txns, key=lambda x: x.trade_date):
            w.writerow([
                t.trade_date, t.settle_date, t.broker, t.txn_type.value,
                t.symbol, t.isin or "", t.description,
                t.country_code or "", t.domicile.value, t.asset_class.value,
                _fmt(t.quantity), _fmt(t.price), t.price_currency or "",
                t.orig_currency, _fmt(t.orig_amount),
                _fmt(t.commission), t.commission_currency or "",
                _fmt(t.wht_rate_actual), _fmt(t.wht_amount_orig),
                _fmt(t.fx_rate_to_eur), _fmt(t.eur_amount),
                _fmt(t.eur_commission), _fmt(t.eur_wht),
                t.notes,
            ])


def _fmt(v: Optional[Decimal]) -> str:
    if v is None:
        return ""
    return str(v.normalize())


# ── Tax Summary TXT ───────────────────────────────────────────────────────────

def _write_tax_summary(s: TaxSummary, path: Path) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [
        "=" * 62,
        f"  KAPITALERTRAG — E1kv KENNZIFFERN",
        f"  Person : {s.person_label}",
        f"  Year   : {s.tax_year}",
        f"  Created: {now}",
        "=" * 62,
        "",
        "  ENTER THESE VALUES INTO FINANZONLINE / E1kv FORM:",
        "",
        f"  ── 1.3.1 Dividenden + Zinsen ─────────────────────────────",
        f"  KZ 862   Inländ. Dividendenerträge          EUR {s.kz_862:>12,.2f}",
        f"  KZ 863   Ausländ. Dividendenerträge          EUR {s.kz_863:>12,.2f}",
        f"  ── 1.3.2 Kursgewinne / Kursverluste ──────────────────────",
        f"  KZ 981   Inländ. Kursgewinne (27,5%)        EUR {s.kz_981:>12,.2f}",
        f"  KZ 994   Ausländ. Kursgewinne (27,5%)        EUR {s.kz_994:>12,.2f}",
        f"  KZ 864   Inländ. Kursgewinne (25%)          EUR {s.kz_864:>12,.2f}",
        f"  KZ 865   Ausländ. Kursgewinne (25%)          EUR {s.kz_865:>12,.2f}",
        f"  KZ 891   Inländ. Kursverluste                EUR {s.kz_891:>12,.2f}",
        f"  KZ 892   Ausländ. Kursverluste               EUR {s.kz_892:>12,.2f}",
        f"  ── 1.3.4 Investmentfonds ─────────────────────────────────",
        f"  KZ 897   Inländ. Fonds-Ausschüttungen       EUR {s.kz_897:>12,.2f}",
        f"  KZ 898   Ausländ. Fonds-Ausschüttungen       EUR {s.kz_898:>12,.2f}",
        f"  KZ 936   Ausschüttungsgleiche Ertr. (Inl.)  EUR {s.kz_936:>12,.2f}",
        f"  KZ 937   Ausschüttungsgleiche Ertr. (Ausl.)  EUR {s.kz_937:>12,.2f}",
        f"  ── Saldo aus Punkt 1.3 ───────────────────────────────────",
        f"  Saldo Inländisch                             EUR {s.saldo_inland:>12,.2f}",
        f"  Saldo Ausländisch                            EUR {s.saldo_ausland:>12,.2f}",
        f"  ── 1.4 / 1.6 Bereits bezahlte Steuer ────────────────────",
        f"  KZ 899   KESt inländ. WP im Ausland          EUR {s.kz_899:>12,.2f}",
        f"  KZ 984   QSt inländ. Einkünfte (27,5%)      EUR {s.kz_984:>12,.2f}",
        f"  KZ 998   QSt ausländ. Einkünfte (27,5%)      EUR {s.kz_998:>12,.2f}",
        "",
        "─" * 62,
        f"  Net taxable income                           EUR {s.net_taxable_eur:>12,.2f}",
        f"  KESt due @ 27.5%                             EUR {s.kest_due_eur:>12,.2f}",
        f"  WHT creditable (max 15%)                     EUR {s.wht_creditable_eur:>12,.2f}",
        f"  KESt remaining to pay                        EUR {s.kest_remaining_eur:>12,.2f}",
        "─" * 62,
        "",
        f"  Transactions processed : {s.transaction_count}",
        f"  Unmatched sells        : {s.unmatched_sells}",
        f"  Missing FX rates       : {s.missing_fx_count}",
        "",
    ]

    if s.warnings:
        lines.append(f"  WARNINGS ({len(s.warnings)}):")
        for w in s.warnings:
            lines.append(f"    • {w}")
        lines.append("")

    lines += [
        "  NOTE: KZ 937 (Ausschüttungsgleiche Erträge) requires OeKB",
        "  fund data and is NOT automatically calculated. Check your",
        "  broker's tax report or oekb.at for any ETF/fund holdings.",
        "",
        "  This output is informational. Verify with your tax consultant.",
        "=" * 62,
    ]

    path.write_text("\n".join(lines), encoding="utf-8")


# ── Excel Dashboard ───────────────────────────────────────────────────────────

def _write_excel(txns: list[NormalizedTransaction],
                 summary: TaxSummary, path: Path,
                 config: dict | None = None,
                 history: list[dict] | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()

    # ── Tab 1: E1kv Summary — built last so formula refs to other tabs work ────
    # (tabs must exist before we write cross-sheet formulas; Dividends/Trades created below)
    ws = wb.active
    ws.title = "E1kv Summary"
    _summary_ws = ws  # filled after other tabs exist

    # ── Tab 2: Year-over-year overview (Verlustausgleich) ─────────────────────
    wo = wb.create_sheet("Overview")
    _fill_overview_sheet(wo, history or [], summary.tax_year, config)

    # ── Tab 3: All Transactions ───────────────────────────────────────────────
    wt = wb.create_sheet("Transactions")
    _fill_transactions_sheet(wt, txns, summary.tax_year)

    # ── Tab 4: Dividends only ─────────────────────────────────────────────────
    wd = wb.create_sheet("Dividends")
    div_txns = [t for t in txns
                if t.txn_type == TransactionType.DIVIDEND
                and t.trade_date.year == summary.tax_year]
    div_refs = _fill_transactions_sheet(wd, div_txns, summary.tax_year,
                                        title="Dividends", show_div_totals=True)

    # ── Tab 5: Trades only ────────────────────────────────────────────────────
    wtr = wb.create_sheet("Trades")
    trade_txns = [t for t in txns
                  if t.txn_type in (TransactionType.BUY, TransactionType.SELL)
                  and t.trade_date.year == summary.tax_year]
    trade_refs = _fill_transactions_sheet(wtr, trade_txns, summary.tax_year,
                                          title="Trades", show_gain_loss=True)

    # ── Fill E1kv Summary now that Dividends/Trades tabs exist ──────────────
    _fill_summary_sheet(_summary_ws, summary, trade_refs=trade_refs, div_refs=div_refs)

    # ── Tab 6: Freedom (static snapshot at config assumptions) ───────────────
    wf = wb.create_sheet("Freedom")
    _fill_freedom_sheet(wf, div_txns, summary, config or {})

    # ── Tab 7: Nichtmeldefonds (only if positions exist) ──────────────────────
    if summary.nichtmeldefonds:
        wnmf = wb.create_sheet("Nichtmeldefonds")
        _fill_nichtmeldefonds_sheet(wnmf, summary)

    # ── Tab 8: Meldefonds (only if positions exist) ───────────────────────────
    if summary.meldefonds:
        wmf = wb.create_sheet("Meldefonds")
        _fill_meldefonds_sheet(wmf, summary)

    wb.save(path)


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL   = "1F3864"   # dark navy — main title
SECTION_FILL  = "FFC000"   # amber — E1kv section headers (1.3.1, 1.3.2 …)
SUBSECT_FILL  = "FFE090"   # light amber — sub-labels within a section
ACCENT_FILL   = "2E75B6"   # mid blue — summary headers
LIGHT_FILL    = "DEEAF1"   # pale blue
SALDO_FILL    = "D6E4F7"   # light blue-grey — Saldo row
WARN_FILL     = "FFF2CC"   # yellow
GREEN_FILL    = "E2EFDA"
RED_FILL      = "FCE4D6"
WHITE         = "FFFFFF"


def _hfill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")


def _fill_summary_sheet(ws, s: TaxSummary,
                        trade_refs: dict | None = None,
                        div_refs: dict | None = None) -> None:
    # Column layout: A=section, B=KZ, C=description, D=Inländisch, E=Ausländisch
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = [1]  # mutable counter
    _kz_rows: dict = {}  # kz_code → row number (for Saldo formula)

    def _r():
        r = row[0]
        row[0] += 1
        return r

    def _ref(refs, key, tab, fallback, negate=False):
        """Return cross-sheet formula if refs available, else Python fallback value."""
        if refs and key in refs:
            sign = "-" if negate else ""
            return f"={sign}{tab}!{refs[key]}"
        return fallback

    def title(text):
        r = _r()
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = _font(bold=True, color=WHITE, size=12)
        c.fill = _hfill(HEADER_FILL)
        c.alignment = _center()
        ws.row_dimensions[r].height = 24

    def col_headers():
        r = _r()
        for col, txt in [("A", ""), ("B", "KZ"), ("C", ""), ("D", "Inländisch EUR"), ("E", "Ausländisch EUR")]:
            c = ws[f"{col}{r}"]
            c.value = txt
            c.font = _font(bold=True, color=WHITE, size=10)
            c.fill = _hfill(ACCENT_FILL)
            c.alignment = _center()
            c.border = _border()

    def section(code, text):
        r = _r()
        ws[f"A{r}"] = code
        ws[f"A{r}"].font = _font(bold=True, size=10)
        ws[f"A{r}"].fill = _hfill(SECTION_FILL)
        ws[f"A{r}"].alignment = _center()
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]
        c.value = text
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(SECTION_FILL)
        ws.row_dimensions[r].height = 18

    def kz_row(kz_in, kz_out, label, val_in, val_out, fill=None, warn=False):
        r = _r()
        if kz_in:
            _kz_rows[kz_in] = r
        if kz_out:
            _kz_rows[kz_out] = r
        ws[f"A{r}"].fill = _hfill(fill or WHITE)
        ws[f"B{r}"] = f"{kz_in}/{kz_out}" if kz_in and kz_out else (kz_in or kz_out or "")
        ws[f"B{r}"].font = _font(bold=True, size=10)
        ws[f"B{r}"].alignment = _center()
        ws[f"B{r}"].fill = _hfill(fill or WHITE)
        ws[f"C{r}"] = label
        ws[f"C{r}"].font = _font(size=10, color="C00000" if warn else "000000")
        ws[f"C{r}"].fill = _hfill(WARN_FILL if warn else (fill or WHITE))
        for col, val in [("D", val_in), ("E", val_out)]:
            c = ws[f"{col}{r}"]
            if val is not None:
                is_formula = isinstance(val, str) and val.startswith("=")
                c.value = val if is_formula else float(val)
                c.number_format = '#,##0.00'
                if not is_formula and float(val) < 0:
                    c.font = Font(color="C00000", size=10)
            c.alignment = _right()
            c.fill = _hfill(WARN_FILL if warn else (fill or WHITE))
        for col in "ABCDE":
            ws[f"{col}{r}"].border = _border()

    def saldo_row(val_in, val_out):
        r = _r()
        _kz_rows["_saldo"] = r
        ws.merge_cells(f"A{r}:C{r}")
        c = ws[f"A{r}"]
        c.value = "Saldo aus Punkt 1.3"
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(SALDO_FILL)
        c.alignment = _center()
        for col, val in [("D", val_in), ("E", val_out)]:
            cell = ws[f"{col}{r}"]
            is_formula = isinstance(val, str) and val.startswith("=")
            cell.value = val if is_formula else float(val)
            cell.number_format = '#,##0.00'
            cell.alignment = _right()
            cell.font = _font(bold=True, size=10)
            if not is_formula:
                cell.font = _font(bold=True, size=10,
                                  color="C00000" if float(val) < 0 else "000000")
            cell.fill = _hfill(SALDO_FILL)
            cell.border = _border()
        for col in "ABC":
            ws[f"{col}{r}"].border = _border()
        ws.row_dimensions[r].height = 18

    def summary_row(label, val_in, val_out, fill=None):
        r = _r()
        ws.merge_cells(f"A{r}:C{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = _font(bold=True, size=10)
        c.fill = _hfill(fill or LIGHT_FILL)
        c.alignment = _center()
        for col, val in [("D", val_in), ("E", val_out)]:
            cell = ws[f"{col}{r}"]
            if val is not None:
                is_formula = isinstance(val, str) and val.startswith("=")
                cell.value = val if is_formula else float(val)
                cell.number_format = '#,##0.00'
            cell.alignment = _right()
            cell.font = _font(bold=True, size=10)
            cell.fill = _hfill(fill or LIGHT_FILL)
            cell.border = _border()
        for col in "ABC":
            ws[f"{col}{r}"].border = _border()
        ws.row_dimensions[r].height = 18

    def blank():
        r = _r()
        ws.row_dimensions[r].height = 5

    # ── Title + column headers ────────────────────────────────────────────────
    title(f"E1kv — Kapitalvermögen  |  {s.person_label}  |  {s.tax_year}")
    col_headers()

    # ── 1.3.1 Dividenden + Zinsen ─────────────────────────────────────────────
    section("1.3.1", "Einkünfte aus der Überlassung von Kapital "
            "(§ 27 Abs. 2 — Dividenden, Zinsen aus Wertpapieren, 27,5%)")
    kz_row("862", "863", "Dividendenerträge + Zinsen",
           _ref(div_refs, "dom_divs", "Dividends", s.kz_862),
           _ref(div_refs, "fgn_divs", "Dividends", s.kz_863))
    blank()

    # ── 1.3.2 Kursgewinne ─────────────────────────────────────────────────────
    section("1.3.2", "Einkünfte aus realisierten Wertsteigerungen von Kapitalvermögen (§ 27 Abs. 3)")
    kz_row("981", "994", "Überschüsse — besonderer Steuersatz 27,5%",
           _ref(trade_refs, "dom_gains", "Trades", s.kz_981),
           _ref(trade_refs, "fgn_gains", "Trades", s.kz_994),
           fill=GREEN_FILL if (s.kz_981 + s.kz_994) > 0 else None)
    kz_row("864", "865", "Überschüsse — besonderer Steuersatz 25% (Wertpapiere vor 2011)",
           s.kz_864, s.kz_865)
    kz_row("891", "892", "Verluste",
           _ref(trade_refs, "dom_losses", "Trades", -s.kz_891 if s.kz_891 else None, negate=True),
           _ref(trade_refs, "fgn_losses", "Trades", -s.kz_892 if s.kz_892 else None, negate=True),
           fill=RED_FILL if (s.kz_891 + s.kz_892) > 0 else None)
    blank()

    # ── 1.3.3 Derivate ────────────────────────────────────────────────────────
    section("1.3.3", "Einkünfte aus verbrieften Derivaten (§ 27 Abs. 4)")
    kz_row("982", "993", "Überschüsse — 27,5%", s.kz_982, s.kz_993)
    kz_row("893", "894", "Überschüsse — 25%",   s.kz_893, s.kz_894)
    kz_row("895", "896", "Verluste",
           -s.kz_895 if s.kz_895 else None, -s.kz_896 if s.kz_896 else None,
           fill=RED_FILL if (s.kz_895 + s.kz_896) > 0 else None)
    blank()

    # ── 1.3.4 Investmentfonds ─────────────────────────────────────────────────
    section("1.3.4", "Einkünfte aus Investmentfonds und Immobilieninvestmentfonds (§ 27 Abs. 3+4 InvFG)")
    kz_row("897", "898", "Ausschüttungen — 27,5%", s.kz_897, s.kz_898)
    has_mf = bool(s.meldefonds)
    kz_row("936", "937",
           "Ausschüttungsgleiche Erträge — 27,5%" + (" ⚠ PLACEHOLDERs in Dataset!" if has_mf and s.kz_937 == 0 else ""),
           s.kz_936, s.kz_937, warn=(not has_mf and s.kz_937 == 0))
    blank()

    # ── 1.3.5 Kryptowährungen ─────────────────────────────────────────────────
    section("1.3.5", "Einkünfte aus Kryptowährungen (§ 27b)")
    kz_row("171",  "",    "Laufende Einkünfte (Mining, Staking)", s.kz_171, None)
    kz_row("173",  "",    "Überschüsse aus Wertsteigerungen",     s.kz_173, None)
    kz_row("175",  "",    "Verluste",
           -s.kz_175 if s.kz_175 else None, None,
           fill=RED_FILL if s.kz_175 > 0 else None)
    blank()

    # ── Saldo 1.3 — formula sums the KZ cells above by row reference ─────────
    # Build formula: sum all inland KZ cells in col D (gains positive, losses already negative)
    def _saldo_formula(col: str, kz_codes: list[str]) -> str:
        refs = [f"{col}{_kz_rows[k]}" for k in kz_codes if k in _kz_rows]
        return ("=" + "+".join(refs)) if refs else "=0"

    inland_kzs  = ["862", "897", "936", "981", "864", "982", "893", "171", "173",
                   "891", "895", "175"]
    ausland_kzs = ["863", "898", "937", "994", "865", "993", "894", "892", "896"]
    saldo_row(_saldo_formula("D", inland_kzs), _saldo_formula("E", ausland_kzs))
    blank()

    # ── 1.4 KESt bereits bezahlt ──────────────────────────────────────────────
    section("1.4", "Kapitalertragssteuer, soweit sie auf die inländischen Kapitaleinküfte entfällt")
    kz_row("899", "", "KESt für inländ. WP, die im Ausland gehalten werden",
           s.kz_899, None, fill=LIGHT_FILL)
    blank()

    # ── 1.5 Abgeltungssteuer ──────────────────────────────────────────────────
    section("1.5", "Abgeltungssteuer nach den Steuerabkommen mit Liechtenstein")
    kz_row("942", "", "Abgeltungssteuer Liechtenstein", s.kz_942, None)
    blank()

    # ── 1.6 Anrechenbare QSt 27.5% ───────────────────────────────────────────
    section("1.6", "Anzurechnende ausländ. (Quellen)Steuer auf Einkünfte — Steuersatz 27,5%")
    kz_row("984", "998", "Anzurechnende Quellensteuer",
           s.kz_984, s.kz_998, fill=GREEN_FILL if s.kz_998 > 0 else None)
    blank()

    # ── 1.7 Anrechenbare QSt 25% ─────────────────────────────────────────────
    section("1.7", "Anzurechnende ausländ. (Quellen)Steuer auf Einkünfte — Steuersatz 25%")
    kz_row("900", "901", "Anzurechnende Quellensteuer 25%", s.kz_900, s.kz_901)
    blank()

    # ── Calculation summary ───────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "BERECHNUNG"
    c.font = _font(bold=True, color=WHITE, size=11)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 20

    # Net taxable = Saldo inland + Saldo ausland → formula if Saldo row was tracked
    saldo_r = _kz_rows.get("_saldo")
    net_formula = f"=D{saldo_r}+E{saldo_r}" if saldo_r else s.saldo_inland + s.saldo_ausland
    summary_row("Steuerpflichtiger Gesamtbetrag (Saldo 1.3 gesamt)", net_formula, None, LIGHT_FILL)
    summary_row("KESt (27,5%)", s.kest_due_eur, None)
    summary_row("Anzurechnende Quellensteuer (KZ 998)", s.wht_creditable_eur, None, GREEN_FILL)
    summary_row("Verbleibende KESt zu bezahlen", s.kest_remaining_eur, None, WARN_FILL)
    blank()

    # ── Notes ─────────────────────────────────────────────────────────────────
    notes_row = row[0]
    using_formulas = bool(trade_refs or div_refs)
    notes = [
        "" if s.meldefonds else "⚠  KZ 936/937 (Ausschüttungsgleiche Erträge): Meldefonds unter 'meldefonds:' in config.local.yaml konfigurieren; AE/WA-Daten in data/oekb_ae.yaml eintragen.",
        "⚠  Nichtmeldefonds (REITs, BDCs): pauschal-AE-Berechnung läuft automatisch — siehe 'Nichtmeldefonds'-Tab.",
        "    Verluste (KZ 891/892) werden hier als negative Werte dargestellt; in FinanzOnline als Absolutbeträge eintragen.",
        "    KeSt-Berechnung (WHT-Anrechnung) enthält Treaty-Rate-Logik — nicht in Excel repliziert; Wert bleibt fix."
        if using_formulas else "",
        "    Diese Ausgabe ist informativ. Bitte mit Steuerberater:in abstimmen.",
    ]
    for i, note in enumerate(n for n in notes if n):
        r = row[0] + i
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = note
        c.font = Font(color="666666", italic=True, size=9)
        ws.row_dimensions[r].height = 14


def _fill_transactions_sheet(ws, txns: list[NormalizedTransaction],
                               year: int, title: str = "Transactions",
                               show_gain_loss: bool = False,
                               show_div_totals: bool = False) -> dict:
    """
    Fill a transactions worksheet.  Returns a dict of cell addresses for
    key summary cells so the E1kv Summary tab can reference them directly.

    Keys (when applicable):
      show_gain_loss=True → "dom_gains", "fgn_gains", "dom_losses", "fgn_losses"
      show_div_totals=True → "dom_divs", "fgn_divs"
    """
    base_headers = [
        "Date", "Type", "Symbol", "ISIN", "Description",
        "Country", "Domicile", "Qty", "Currency",
        "Orig Amount", "Commission", "WHT (orig)",
        "FX Rate", "EUR Amount", "EUR Commission", "EUR WHT",
        "Notes",
    ]
    gl_headers = ["Gain/Loss EUR", "Cost Basis EUR"] if show_gain_loss else []
    headers = base_headers + gl_headers

    # Column indices (1-based)
    COL_TYPE     = 2   # B — transaction type ("buy"/"sell"/"dividend")
    COL_DOMICILE = 7   # G — domicile ("domestic"/"foreign")
    COL_EUR_AMT  = 14  # N — EUR amount (used for dividend totals)
    COL_EUR_WHT  = 16  # P — EUR WHT (used for WHT totals)
    COL_GL       = len(base_headers) + 1  # R (18) — Gain/Loss EUR
    COL_COST     = len(base_headers) + 2  # S (19) — Cost Basis EUR

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _hfill(HEADER_FILL)
        cell.alignment = _center()
        cell.border = _border()

    txns_sorted = sorted(txns, key=lambda t: t.trade_date)
    for t in txns_sorted:
        row_data = [
            t.trade_date.isoformat(),
            t.txn_type.value,
            t.symbol,
            t.isin or "",
            t.description[:60],
            t.country_code or "",
            t.domicile.value,
            float(t.quantity) if t.quantity else "",
            t.orig_currency,
            float(t.orig_amount),
            float(t.commission) if t.commission else "",
            float(t.wht_amount_orig) if t.wht_amount_orig else "",
            float(t.fx_rate_to_eur) if t.fx_rate_to_eur else "",
            float(t.eur_amount) if t.eur_amount else "",
            float(t.eur_commission) if t.eur_commission else "",
            float(t.eur_wht) if t.eur_wht else "",
            t.notes,
        ]
        if show_gain_loss:
            row_data.append(float(t.eur_gain_loss) if t.eur_gain_loss is not None else "")
            row_data.append(float(t.eur_cost_basis) if t.eur_cost_basis is not None else "")
        ws.append(row_data)

    data_last = ws.max_row   # last data row (before any summary rows)

    # ── Number format and conditional colour for gain/loss ────────────────────
    eur_cols = [10, 11, 12, 14, 15, 16]
    for row_idx in range(2, data_last + 1):
        for col_idx in eur_cols:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00'
        if show_gain_loss:
            gl_val = ws.cell(row_idx, COL_GL).value
            if isinstance(gl_val, (int, float)):
                ws.cell(row_idx, COL_GL).number_format = '#,##0.00'
                if gl_val > 0.005:
                    ws.cell(row_idx, COL_GL).fill = _hfill(GREEN_FILL)
                elif gl_val < -0.005:
                    ws.cell(row_idx, COL_GL).fill = _hfill(RED_FILL)
                    ws.cell(row_idx, COL_GL).font = Font(color="C00000", size=10)
            cb_val = ws.cell(row_idx, COL_COST).value
            if isinstance(cb_val, (int, float)):
                ws.cell(row_idx, COL_COST).number_format = '#,##0.00'

    # ── Summary rows ──────────────────────────────────────────────────────────
    cell_refs: dict = {}

    def _sum_hdr(row_idx: int, label: str, fill: str) -> None:
        for col in range(1, len(headers) + 1):
            ws.cell(row_idx, col).fill = _hfill(fill)
        ws.merge_cells(f"A{row_idx}:{get_column_letter(len(headers) - 1)}{row_idx}")
        c = ws.cell(row_idx, 1)
        c.value = label
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _hfill(ACCENT_FILL)
        c.alignment = _center()
        ws.row_dimensions[row_idx].height = 16

    def _sum_row(row_idx: int, label: str, formula: str,
                 val_col: int, lbl_fill: str, val_fill: str,
                 negative: bool = False) -> str:
        """Write a labelled formula row; return the value cell address."""
        ws.merge_cells(f"A{row_idx}:{get_column_letter(val_col - 1)}{row_idx}")
        lbl = ws.cell(row_idx, 1)
        lbl.value = label
        lbl.font = _font(bold=True, size=10)
        lbl.fill = _hfill(lbl_fill)
        lbl.alignment = _right()
        lbl.border = _border()
        cell = ws.cell(row_idx, val_col)
        cell.value = formula
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, color="C00000" if negative else "000000", size=10)
        cell.fill = _hfill(val_fill)
        cell.alignment = _right()
        cell.border = _border()
        ws.row_dimensions[row_idx].height = 15
        return f"{get_column_letter(val_col)}{row_idx}"

    def _rng(col_letter: str) -> str:
        """Absolute range string for a column over all data rows."""
        return f"${col_letter}$2:${col_letter}${data_last}"

    if show_gain_loss and data_last >= 2:
        gl_L  = get_column_letter(COL_GL)
        dom_L = get_column_letter(COL_DOMICILE)
        typ_L = get_column_letter(COL_TYPE)

        hdr_r = data_last + 1
        _sum_hdr(hdr_r, "GAIN / LOSS SUMMARY  (SELL transactions, tax year)", ACCENT_FILL)

        cell_refs["dom_gains"] = _sum_row(
            data_last + 2,
            "Domestic Gains  (→ KZ 981)",
            f"=SUMPRODUCT(({_rng(typ_L)}=\"sell\")*({_rng(dom_L)}=\"domestic\")*({_rng(gl_L)}>0),{_rng(gl_L)})",
            COL_GL, GREEN_FILL, GREEN_FILL,
        )
        cell_refs["fgn_gains"] = _sum_row(
            data_last + 3,
            "Foreign Gains  (→ KZ 994)",
            f"=SUMPRODUCT(({_rng(typ_L)}=\"sell\")*({_rng(dom_L)}=\"foreign\")*({_rng(gl_L)}>0),{_rng(gl_L)})",
            COL_GL, GREEN_FILL, GREEN_FILL,
        )
        cell_refs["dom_losses"] = _sum_row(
            data_last + 4,
            "Domestic Losses  (→ KZ 891, positive)",
            f"=SUMPRODUCT(({_rng(typ_L)}=\"sell\")*({_rng(dom_L)}=\"domestic\")*({_rng(gl_L)}<0),-{_rng(gl_L)})",
            COL_GL, RED_FILL, RED_FILL, negative=True,
        )
        cell_refs["fgn_losses"] = _sum_row(
            data_last + 5,
            "Foreign Losses  (→ KZ 892, positive)",
            f"=SUMPRODUCT(({_rng(typ_L)}=\"sell\")*({_rng(dom_L)}=\"foreign\")*({_rng(gl_L)}<0),-{_rng(gl_L)})",
            COL_GL, RED_FILL, RED_FILL, negative=True,
        )
        _sum_row(
            data_last + 6,
            "Net Gain/Loss (all)",
            f"={gl_L}{data_last + 2}-{gl_L}{data_last + 4}+{gl_L}{data_last + 3}-{gl_L}{data_last + 5}",
            COL_GL, LIGHT_FILL, LIGHT_FILL,
        )

    if show_div_totals and data_last >= 2:
        eur_L = get_column_letter(COL_EUR_AMT)
        wht_L = get_column_letter(COL_EUR_WHT)
        dom_L = get_column_letter(COL_DOMICILE)

        hdr_r = data_last + 1
        _sum_hdr(hdr_r, "DIVIDEND SUMMARY  (tax year)", ACCENT_FILL)

        cell_refs["dom_divs"] = _sum_row(
            data_last + 2,
            "Domestic Dividends  (→ KZ 862)",
            f"=SUMIF({_rng(dom_L)},\"domestic\",{_rng(eur_L)})",
            COL_EUR_AMT, GREEN_FILL, GREEN_FILL,
        )
        cell_refs["fgn_divs"] = _sum_row(
            data_last + 3,
            "Foreign Dividends  (→ KZ 863)",
            f"=SUMIF({_rng(dom_L)},\"foreign\",{_rng(eur_L)})",
            COL_EUR_AMT, GREEN_FILL, GREEN_FILL,
        )
        _sum_row(
            data_last + 4,
            "Total Dividends",
            f"={eur_L}{data_last + 2}+{eur_L}{data_last + 3}",
            COL_EUR_AMT, LIGHT_FILL, LIGHT_FILL,
        )
        _sum_row(
            data_last + 5,
            "Total WHT paid (EUR)",
            f"=SUM({_rng(wht_L)})",
            COL_EUR_WHT, WARN_FILL, WARN_FILL,
        )

    # ── Auto-width columns ────────────────────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row_idx, col_idx).value or ""))
             for row_idx in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    return cell_refs


# ── Overview tab (Verlustausgleich) ──────────────────────────────────────────

_OV_COLS = [
    ("Year",               8),
    ("Dividends",         14),
    ("Dom Gains\nKZ 981", 14),
    ("Dom Losses\nKZ 891",14),
    ("Fgn Gains\nKZ 994", 14),
    ("Fgn Losses\nKZ 892",14),
    ("Net Taxable",       14),
    ("KeSt 27.5%",        14),
    ("WHT Credited",      14),
    ("KeSt Remaining",    15),
    ("Div YoY %",         10),
]


def _fill_overview_sheet(ws, history: list[dict], current_year: int,
                         config: dict | None = None) -> None:
    # Column widths
    for i, (_, w) in enumerate(_OV_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    person = history[0]["person_label"] if history else ""
    years  = [str(e.get("tax_year", "")) for e in history]
    yr_range = f"{years[0]}–{years[-1]}" if len(years) > 1 else (years[0] if years else "")

    row = [1]

    def _r():
        r = row[0]; row[0] += 1; return r

    # Title
    r = _r()
    ncols = len(_OV_COLS)
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    c = ws[f"A{r}"]
    c.value = f"Verlustausgleich Übersicht  |  {person}  |  {yr_range}"
    c.font = _font(bold=True, color=WHITE, size=12)
    c.fill = _hfill(HEADER_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 24

    # Header row (two-line headers for KZ columns)
    r = _r()
    for col_idx, (label, _) in enumerate(_OV_COLS, 1):
        c = ws.cell(r, col_idx, label)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _hfill(ACCENT_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[r].height = 28
    # Supplementary chart data headers (cols 20-22, 25-27) — chart-only, no styling
    ws.cell(r, 20, "Interest EUR")
    ws.cell(r, 21, "Transactions")
    ws.cell(r, 22, "FIRE %")
    ws.cell(r, 25, "Month")
    ws.cell(r, 26, "Div/Month EUR")
    ws.cell(r, 27, "Tx/Month")
    data_start_row = row[0]  # first data row (after title + header)

    def _d(entry: dict, key: str) -> float:
        try:
            return float(Decimal(entry.get(key, "0") or "0"))
        except Exception:
            return 0.0

    # Data rows — use per-KZ fields for domestic/foreign split
    totals = [0.0] * (len(_OV_COLS) - 1)  # skip Year column
    for i, entry in enumerate(history):
        yr         = entry.get("tax_year")
        divs       = _d(entry, "total_dividends_eur")
        dom_gains  = _d(entry, "kz_981")
        dom_losses = _d(entry, "kz_891")
        fgn_gains  = _d(entry, "kz_994")
        fgn_losses = _d(entry, "kz_892")
        net        = _d(entry, "net_taxable_eur")
        kest       = _d(entry, "kest_due_eur")
        wht        = _d(entry, "wht_creditable_eur")
        remain     = _d(entry, "kest_remaining_eur")

        is_current = (yr == current_year)
        row_fill = LIGHT_FILL if is_current else WHITE

        r = _r()
        c = ws.cell(r, 1, yr)
        c.font = _font(bold=is_current, size=10)
        c.fill = _hfill(row_fill)
        c.alignment = _center()
        c.border = _border()

        for col_idx, val, color in [
            (2,  divs,        None),
            (3,  dom_gains,   GREEN_FILL if dom_gains > 0 else None),
            (4,  -dom_losses, RED_FILL if dom_losses > 0 else None),
            (5,  fgn_gains,   GREEN_FILL if fgn_gains > 0 else None),
            (6,  -fgn_losses, RED_FILL if fgn_losses > 0 else None),
            (7,  net,         None),
            (8,  kest,        None),
            (9,  wht,         GREEN_FILL if wht > 0 else None),
            (10, remain,      WARN_FILL if remain > 0.01 else GREEN_FILL),
        ]:
            cell = ws.cell(r, col_idx, val)
            cell.number_format = '#,##0.00'
            cell.alignment = _right()
            cell.font = _font(bold=is_current, size=10,
                              color="C00000" if val < -0.005 else "000000")
            cell.fill = _hfill(color or (LIGHT_FILL if is_current else WHITE))
            cell.border = _border()

        # Div YoY % (col 11)
        if i > 0:
            prev_divs = _d(history[i - 1], "total_dividends_eur")
            yoy = ((divs - prev_divs) / prev_divs * 100) if prev_divs > 0 else 0.0
            yoy_txt   = f"{yoy:+.1f}%"
            yoy_color = "008000" if yoy > 0.05 else ("C00000" if yoy < -0.05 else "000000")
        else:
            yoy_txt   = "—"
            yoy_color = "888888"
        c11 = ws.cell(r, 11, yoy_txt)
        c11.font      = _font(bold=is_current, size=10, color=yoy_color)
        c11.fill      = _hfill(row_fill)
        c11.alignment = _center()
        c11.border    = _border()

        # Supplementary data for charts (cols 20-22, hidden area)
        ws.cell(r, 20, _d(entry, "interest_eur"))
        ws.cell(r, 21, int(entry.get("transaction_count") or 0))
        if config:
            _mexp = float(
                (config.get("freedom_dashboard") or {}).get("monthly_expenses_eur", 0) or 0
            )
            ws.cell(r, 22, round(divs / (_mexp * 12) * 100, 1) if _mexp > 0 else 0.0)

        totals[0] += divs
        totals[1] += dom_gains
        totals[2] += dom_losses
        totals[3] += fgn_gains
        totals[4] += fgn_losses
        totals[5] += net
        totals[6] += kest
        totals[7] += wht
        totals[8] += remain

        ws.row_dimensions[r].height = 16

    if not history:
        return

    # Totals row (losses shown as negative in the totals)
    r = _r()
    c = ws.cell(r, 1, "TOTAL")
    c.font = _font(bold=True, size=10, color=WHITE)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    c.border = _border()

    signed_totals = [
        totals[0],           # divs
        totals[1],           # dom gains
        -totals[2],          # dom losses (negative)
        totals[3],           # fgn gains
        -totals[4],          # fgn losses (negative)
        totals[5],           # net taxable
        totals[6],           # kest
        totals[7],           # wht
        totals[8],           # remaining
        "",                  # Div YoY % — not aggregated
    ]
    for col_idx, val in enumerate(signed_totals, 2):
        cell = ws.cell(r, col_idx, val if val != "" else None)
        if isinstance(val, float):
            cell.number_format = '#,##0.00'
        cell.alignment = _right()
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _hfill(ACCENT_FILL)
        cell.border = _border()
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"

    if len(history) < 2:
        return

    data_end_row = r - 1  # last data row before TOTAL
    cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    chart_top = r + 2
    CW, CH = 15, 9
    CROW   = 18

    # Monthly mini-table (cols 25-27, rows data_start_row to data_start_row+11)
    monthly_entry = next(
        (e for e in reversed(history) if e.get("monthly_dividends")),
        None,
    )
    monthly_year = (monthly_entry.get("tax_year", current_year)
                    if monthly_entry else current_year)
    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m_idx, m_label in enumerate(month_labels, 1):
        key  = f"{m_idx:02d}"
        mrow = data_start_row + m_idx - 1
        ws.cell(mrow, 25, m_label)
        m_div = float(Decimal(str(
            ((monthly_entry or {}).get("monthly_dividends") or {}).get(key) or "0"
        ))) if monthly_entry else 0.0
        m_tx = int(
            (((monthly_entry or {}).get("monthly_transaction_counts") or {}).get(key) or 0)
        )
        ws.cell(mrow, 26, m_div)
        ws.cell(mrow, 27, m_tx)
    month_end_row = data_start_row + 11

    def _col_chart(title: str, y_title: str = "EUR") -> BarChart:
        ch = BarChart()
        ch.type = "col"
        ch.grouping = "clustered"
        ch.title = title
        ch.y_axis.title = y_title
        ch.width = CW
        ch.height = CH
        return ch

    # Chart 1 (ROW 0, LEFT): Dividends per Year
    c1 = _col_chart("Dividend Income by Year")
    c1.add_data(Reference(ws, min_col=2, min_row=2, max_row=data_end_row), titles_from_data=True)
    c1.set_categories(cats)
    ws.add_chart(c1, f"A{chart_top}")

    # Chart 2 (ROW 0, RIGHT): KeSt Remaining per Year
    c2 = _col_chart("KeSt Remaining by Year")
    c2.add_data(Reference(ws, min_col=10, min_row=2, max_row=data_end_row), titles_from_data=True)
    c2.set_categories(cats)
    ws.add_chart(c2, f"L{chart_top}")

    # Chart 3 (ROW 1, LEFT): Income Sources stacked (Dividends + Interest)
    c3 = BarChart()
    c3.type = "col"
    c3.grouping = "stacked"
    c3.title = "Income Sources by Year"
    c3.y_axis.title = "EUR"
    c3.width = CW
    c3.height = CH
    c3.add_data(Reference(ws, min_col=2,  min_row=2, max_row=data_end_row), titles_from_data=True)
    c3.add_data(Reference(ws, min_col=20, min_row=2, max_row=data_end_row), titles_from_data=True)
    c3.set_categories(cats)
    ws.add_chart(c3, f"A{chart_top + CROW}")

    # Chart 4 (ROW 1, RIGHT): Transactions per Year
    c4 = _col_chart("Transactions by Year", y_title="Count")
    c4.add_data(Reference(ws, min_col=21, min_row=2, max_row=data_end_row), titles_from_data=True)
    c4.set_categories(cats)
    ws.add_chart(c4, f"L{chart_top + CROW}")

    has_fire = bool(
        config and (config.get("freedom_dashboard") or {}).get("monthly_expenses_eur")
    )
    if has_fire:
        # Chart 5 (ROW 2, LEFT): FIRE Coverage % per Year
        c5 = LineChart()
        c5.title = "FIRE Coverage % by Year"
        c5.y_axis.title = "%"
        c5.width = CW
        c5.height = CH
        c5.add_data(Reference(ws, min_col=22, min_row=2, max_row=data_end_row), titles_from_data=True)
        c5.set_categories(cats)
        ws.add_chart(c5, f"A{chart_top + CROW * 2}")

        # Chart 6 (ROW 2, RIGHT): Dividends per Month
        c6 = _col_chart(f"Dividends by Month ({monthly_year})")
        c6.add_data(Reference(ws, min_col=26, min_row=data_start_row, max_row=month_end_row))
        c6.set_categories(Reference(ws, min_col=25, min_row=data_start_row, max_row=month_end_row))
        ws.add_chart(c6, f"L{chart_top + CROW * 2}")

        # Chart 7 (ROW 3, LEFT): Transactions per Month
        c7 = _col_chart(f"Transactions by Month ({monthly_year})", y_title="Count")
        c7.add_data(Reference(ws, min_col=27, min_row=data_start_row, max_row=month_end_row))
        c7.set_categories(Reference(ws, min_col=25, min_row=data_start_row, max_row=month_end_row))
        ws.add_chart(c7, f"A{chart_top + CROW * 3}")
    else:
        # Chart 5 (ROW 2, LEFT): Dividends per Month
        c5 = _col_chart(f"Dividends by Month ({monthly_year})")
        c5.add_data(Reference(ws, min_col=26, min_row=data_start_row, max_row=month_end_row))
        c5.set_categories(Reference(ws, min_col=25, min_row=data_start_row, max_row=month_end_row))
        ws.add_chart(c5, f"A{chart_top + CROW * 2}")

        # Chart 6 (ROW 2, RIGHT): Transactions per Month
        c6 = _col_chart(f"Transactions by Month ({monthly_year})", y_title="Count")
        c6.add_data(Reference(ws, min_col=27, min_row=data_start_row, max_row=month_end_row))
        c6.set_categories(Reference(ws, min_col=25, min_row=data_start_row, max_row=month_end_row))
        ws.add_chart(c6, f"L{chart_top + CROW * 2}")


# ── Freedom tab ──────────────────────────────────────────────────────────────

_FD_DEFAULTS = {
    "portfolio_eur": 10000,
    "monthly_expenses_eur": 2000,
    "monthly_contribution_eur": 500,
    "yield_pct": 3.0,
    "growth_pct": 7.0,
}

GREEN_DARK  = "375623"
BLUE_LIGHT2 = "BDD7EE"
PROJ_FILL   = "E7F0FC"


def _fill_freedom_sheet(
    ws,
    div_txns: list[NormalizedTransaction],
    summary: TaxSummary,
    config: dict,
) -> None:
    fd = {**_FD_DEFAULTS, **config.get("freedom_dashboard", {})}

    # Use computed portfolio value if available, else fall back to config.
    # Add portfolio_eur_supplement (e.g. SAXO manual estimate) — mirrors freedom.py HTML logic.
    supplement = float(fd.get("portfolio_eur_supplement", 0))
    computed = summary.portfolio_eur_computed
    if computed is not None and computed > ZERO:
        portfolio = float(computed) + supplement
        portfolio_label = "Portfolio Value (computed)" + (" + manual supplement" if supplement else "")
    else:
        portfolio = float(fd["portfolio_eur"])
        portfolio_label = "Portfolio Value (config)"
    monthly_exp   = float(fd["monthly_expenses_eur"])
    monthly_cont  = float(fd["monthly_contribution_eur"])
    # Use computed trailing yield if available, else fall back to config
    if summary.dividend_yield_computed is not None:
        yield_pct       = summary.dividend_yield_computed / 100.0
        yield_pct_label = f"{summary.dividend_yield_computed:.2f}% (computed)"
    else:
        yield_pct       = float(fd["yield_pct"]) / 100.0
        yield_pct_label = f"{fd['yield_pct']}% (config)"
    growth_pct    = float(fd["growth_pct"]) / 100.0

    annual_div    = float(summary.total_dividends_eur)
    monthly_div   = annual_div / 12.0
    freedom_pct   = (monthly_div / monthly_exp * 100.0) if monthly_exp else 0.0

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10

    row = [1]

    def _r():
        r = row[0]; row[0] += 1; return r

    def _cell(r, col, value=None, bold=False, color="000000", size=11,
              fill=None, fmt=None, align="left", border=False):
        c = ws.cell(r, col, value)
        c.font = _font(bold=bold, color=color, size=size)
        if fill:
            c.fill = _hfill(fill)
        if fmt:
            c.number_format = fmt
        c.alignment = Alignment(horizontal=align, vertical="center")
        if border:
            c.border = _border()
        return c

    # ── Title ─────────────────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = f"Financial Freedom  |  {summary.person_label}  |  {summary.tax_year}"
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _hfill(HEADER_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 26

    r = _r()
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  "
               f"Actual dividends {summary.tax_year}  |  "
               f"Projection: {yield_pct_label} yield · {fd['growth_pct']}% growth · "
               f"€{fd['monthly_contribution_eur']:,.0f}/mo contribution")
    c.font = _font(color="555555", size=9)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 14

    row[0] += 1  # blank spacer

    # ── Key Metrics ───────────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "KEY METRICS"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    metrics = [
        ("Annual Dividends (actual)",      annual_div,   '#,##0.00 "EUR"', None),
        ("Monthly Dividends (actual)",      monthly_div,  '#,##0.00 "EUR"', None),
        ("Monthly Expenses (config)",       monthly_exp,  '#,##0.00 "EUR"', None),
        ("Monthly Surplus / Deficit",       monthly_div - monthly_exp,
         '#,##0.00 "EUR"', GREEN_FILL if monthly_div >= monthly_exp else RED_FILL),
        ("Freedom %",                       freedom_pct,  '0.0"%"',          None),
        (portfolio_label,                   portfolio,    '#,##0 "EUR"',     LIGHT_FILL),
    ]
    for label, val, fmt, fill in metrics:
        r = _r()
        _cell(r, 1, label, bold=True, size=10, fill=fill or WHITE, border=True)
        c = _cell(r, 2, val, bold=True, size=10, fill=fill or WHITE,
                  fmt=fmt, align="right", border=True)
        if label == "Freedom %" :
            pct = min(val, 100.0)
            bar_fill = GREEN_FILL if pct >= 100 else (SECTION_FILL if pct >= 50 else LIGHT_FILL)
            for col in range(3, 7):
                ws.cell(r, col).fill = _hfill(bar_fill if col <= 2 + int(pct / 25) + 1 else WHITE)
        ws.row_dimensions[r].height = 16

    row[0] += 1  # blank

    # ── Portfolio Holdings ────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = f"PORTFOLIO HOLDINGS  —  31 Dec {summary.tax_year}"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    # header row
    r = _r()
    for col, txt in [
        (1, "Symbol [Type]"), (2, "Qty"), (3, "EUR Value"),
        (4, "Port%"), (5, "Divs EUR"), (6, "Yield%"),
    ]:
        _cell(r, col, txt, bold=True, color=WHITE, size=9, fill=ACCENT_FILL, align="center", border=True)
    ws.row_dimensions[r].height = 15

    positions = summary.portfolio_positions
    fd_cfg = config.get("freedom_dashboard", {})
    group_by_type = fd_cfg.get("holdings_group_by_type", False)

    if not positions:
        r = _r()
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = "No portfolio positions available — run with full broker history"
        c.font = _font(color="888888", size=9)
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 14
    else:
        last_type = None
        for p in positions:
            # Group header row (if grouping enabled)
            if group_by_type and p.fund_type != last_type:
                last_type = p.fund_type
                r = _r()
                ws.merge_cells(f"A{r}:F{r}")
                c = ws[f"A{r}"]
                c.value = p.fund_type.upper()
                c.font = _font(bold=True, size=9)
                c.fill = _hfill(SECTION_FILL)
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[r].height = 13

            r = _r()
            sym_label = f"{p.symbol} [{p.fund_type}]"
            sym_bold = not p.is_synthetic and p.fund_type != "Sold"
            _cell(r, 1, sym_label, bold=sym_bold, size=9, fill=WHITE, border=True)

            # Qty col
            if p.is_synthetic:
                qty_val = f"~{int(p.qty)}"
                _cell(r, 2, qty_val, size=9, fill=WHITE, align="right", border=True)
                ws.cell(r, 2).font = _font(color="888888", size=9)
            elif p.qty > ZERO:
                _cell(r, 2, int(p.qty), size=9, fill=WHITE, align="right", border=True)
            else:
                _cell(r, 2, "—", size=9, fill=WHITE, align="right", border=True)

            # EUR Value col
            if p.eur_value > ZERO:
                _cell(r, 3, float(p.eur_value), size=9, fill=WHITE, fmt='#,##0', align="right", border=True)
            else:
                _cell(r, 3, "—", size=9, fill=WHITE, align="right", border=True)

            # Port% col
            if p.portfolio_pct is not None:
                _cell(r, 4, p.portfolio_pct, size=9, fill=WHITE, fmt='0.0"%"', align="right", border=True)
            else:
                _cell(r, 4, "—", size=9, fill=WHITE, align="right", border=True)

            # Divs EUR col
            if p.dividends_eur > ZERO:
                _cell(r, 5, float(p.dividends_eur), size=9, fill=WHITE, fmt='#,##0.00', align="right", border=True)
            else:
                _cell(r, 5, "—", size=9, fill=WHITE, align="right", border=True)

            # Yield% col
            if p.yield_pct is not None:
                yld_fill = GREEN_FILL if p.yield_pct > 5 else WHITE
                _cell(r, 6, p.yield_pct, size=9, fill=yld_fill, fmt='0.0"%"', align="right", border=True)
            else:
                _cell(r, 6, "—", size=9, fill=WHITE, align="right", border=True)

            ws.row_dimensions[r].height = 14

        # Totals row
        total_port = sum(float(p.eur_value) for p in positions)
        total_divs = sum(float(p.dividends_eur) for p in positions)
        overall_yield = (total_divs / total_port * 100) if total_port > 0 else None

        r = _r()
        _cell(r, 1, "TOTAL", bold=True, size=9, fill=LIGHT_FILL, border=True)
        _cell(r, 2, "", size=9, fill=LIGHT_FILL, border=True)
        _cell(r, 3, total_port, bold=True, size=9, fill=LIGHT_FILL, fmt='#,##0', align="right", border=True)
        _cell(r, 4, "", size=9, fill=LIGHT_FILL, border=True)
        _cell(r, 5, total_divs, bold=True, size=9, fill=LIGHT_FILL, fmt='#,##0.00', align="right", border=True)
        if overall_yield is not None:
            _cell(r, 6, overall_yield, bold=True, size=9, fill=LIGHT_FILL, fmt='0.0"%"', align="right", border=True)
        else:
            _cell(r, 6, "—", bold=True, size=9, fill=LIGHT_FILL, align="right", border=True)
        ws.row_dimensions[r].height = 14

    row[0] += 1  # blank

    # ── 10-Year Projection ────────────────────────────────────────────────────
    r = _r()
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "10-YEAR PROJECTION  (static at config assumptions)"
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _hfill(ACCENT_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 18

    r = _r()
    for col, txt in [(1, "Year"), (2, "Portfolio EUR"), (3, "Annual Divs EUR"),
                     (4, "Monthly Divs EUR"), (5, "Freedom %")]:
        _cell(r, col, txt, bold=True, color=WHITE, size=9, fill=ACCENT_FILL, align="center", border=True)
    ws.row_dimensions[r].height = 15

    port = portfolio
    for i in range(11):
        yr = summary.tax_year + i
        ann_d = port * yield_pct if i > 0 else annual_div
        mo_d  = ann_d / 12.0
        fp    = mo_d / monthly_exp * 100.0 if monthly_exp else 0.0
        fill  = GREEN_FILL if fp >= 100 else (PROJ_FILL if i % 2 == 0 else WHITE)

        r = _r()
        _cell(r, 1, yr,    bold=(i == 0), size=9, fill=fill, align="center", border=True)
        _cell(r, 2, port,  size=9, fill=fill, fmt='#,##0', align="right", border=True)
        _cell(r, 3, ann_d, size=9, fill=fill, fmt='#,##0.00', align="right", border=True)
        _cell(r, 4, mo_d,  size=9, fill=fill, fmt='#,##0.00', align="right", border=True)
        _cell(r, 5, fp,    bold=(fp >= 100), size=9, fill=fill, fmt='0.0"%"',
              align="right", border=True)
        ws.row_dimensions[r].height = 14

        port = port * (1 + growth_pct) + monthly_cont * 12

    # note
    r = _r() + 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    port_note = (
        "auto-computed from remaining FIFO lots × Dec 31 prices."
        if (summary.portfolio_eur_computed is not None and summary.portfolio_eur_computed > ZERO)
        else "set portfolio_eur in config.local.yaml for accurate projections."
    )
    c.value = (f"Year 0 = actual dividends. Years 1–10: portfolio × {yield_pct_label} yield, "
               f"{fd['growth_pct']}% annual growth, +€{fd['monthly_contribution_eur']:,.0f}/mo contribution. "
               f"Portfolio value: {port_note}")
    c.font = _font(color="888888", size=8)
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes = "A3"


# ── Nichtmeldefonds tab ───────────────────────────────────────────────────────

def _fill_nichtmeldefonds_sheet(ws, summary) -> None:
    from openpyxl.styles import Font as OFont

    ws.column_dimensions["A"].width = 8    # Symbol
    ws.column_dimensions["B"].width = 6    # Type
    ws.column_dimensions["C"].width = 28   # Name
    ws.column_dimensions["D"].width = 5    # Ccy
    ws.column_dimensions["E"].width = 10   # Shares
    ws.column_dimensions["F"].width = 12   # Jan 1 price
    ws.column_dimensions["G"].width = 12   # Dec 31 price
    ws.column_dimensions["H"].width = 11   # Annual gain/sh
    ws.column_dimensions["I"].width = 10   # AE/sh (90%)
    ws.column_dimensions["J"].width = 10   # AE/sh (10%)
    ws.column_dimensions["K"].width = 10   # AE/sh used
    ws.column_dimensions["L"].width = 13   # AE total native
    ws.column_dimensions["M"].width = 11   # FX Dec31
    ws.column_dimensions["N"].width = 13   # AE total EUR
    ws.column_dimensions["O"].width = 13   # KeSt EUR
    ws.column_dimensions["P"].width = 13   # Cost basis +

    headers = [
        "Symbol", "Type", "Name", "Ccy", "Shares",
        "Jan 1 Price", "Dec 31 Price", "Gain/Share",
        "AE 90%/sh", "AE 10%/sh", "AE/share",
        f"AE ({summary.tax_year} native)", "FX Dec31",
        "AE (EUR)", "KeSt 27.5% (EUR)", "Cost Basis +"
    ]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _hfill(SECTION_FILL)
        cell.alignment = _center()
        cell.border = _border()

    # Data rows
    for r in summary.nichtmeldefonds:
        def fv(v):
            return float(v) if v else None

        row_data = [
            r.symbol,
            r.fund_type,
            r.name,
            r.currency,
            float(r.shares_held),
            fv(r.price_jan1),
            fv(r.price_dec31),
            fv(r.annual_gain_per_share),
            fv(r.ae_90pct_per_share),
            fv(r.ae_10pct_per_share),
            fv(r.ae_per_share),
            fv(r.ae_total_native),
            fv(r.fx_dec31),
            fv(r.ae_total_eur),
            fv(r.kest_due_eur),
            fv(r.cost_basis_adj_eur),
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Formatting
        ws.cell(row_idx, 1).font = _font(bold=True)       # symbol bold
        for col_idx in [6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00'
        ws.cell(row_idx, 13).number_format = '0.0000'     # FX rate
        for col_idx in [14, 15, 16]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00 "EUR"'
            ws.cell(row_idx, col_idx).font = _font(bold=(col_idx == 15))
        ws.cell(row_idx, 15).fill = _hfill(WARN_FILL)     # KeSt highlight

        if r.warning:
            ws.cell(row_idx, 1).font = OFont(color="C00000", bold=True)

        for col_idx in range(1, 17):
            ws.cell(row_idx, col_idx).border = _border()

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(total_row, 13).value = "TOTAL"
    ws.cell(total_row, 13).font = _font(bold=True)
    ws.cell(total_row, 14).value = float(summary.nichtmeldefonds_ae_eur)
    ws.cell(total_row, 14).number_format = '#,##0.00 "EUR"'
    ws.cell(total_row, 14).font = _font(bold=True)
    ws.cell(total_row, 14).fill = _hfill(LIGHT_FILL)
    ws.cell(total_row, 15).value = float(summary.nichtmeldefonds_kest_eur)
    ws.cell(total_row, 15).number_format = '#,##0.00 "EUR"'
    ws.cell(total_row, 15).font = _font(bold=True)
    ws.cell(total_row, 15).fill = _hfill(WARN_FILL)
    for col_idx in [14, 15]:
        ws.cell(total_row, col_idx).border = _border()

    # Notes
    note_row = total_row + 2
    notes = [
        f"KZ 937 (Ausschüttungsgleiche Erträge Ausland) = {float(summary.nichtmeldefonds_ae_eur):,.2f} EUR  →  KeSt = {float(summary.nichtmeldefonds_kest_eur):,.2f} EUR",
        "Cost Basis + = Erhöhung der steuerlichen Anschaffungskosten um die AE (verhindert Doppelbesteuerung beim Verkauf).",
        "AE/share = max(90% × Jahresgewinn/Anteil,  10% × Jahresendkurs/Anteil)   per § 186 Abs. 2 InvFG.",
        "Preise (dec31_prices in config) manuell eintragen: Jahresschlusskurs lt. IBKR-Portfolioaufstellung oder Yahoo Finance.",
    ]
    for i, note in enumerate(notes):
        ws.merge_cells(f"A{note_row + i}:P{note_row + i}")
        c = ws.cell(note_row + i, 1)
        c.value = note
        c.font = OFont(color="666666", italic=True, size=9)

    ws.freeze_panes = "A2"


# ── Meldefonds tab ────────────────────────────────────────────────────────────

def _fill_meldefonds_sheet(ws, summary) -> None:
    from openpyxl.styles import Font as OFont

    ws.column_dimensions["A"].width = 16   # ISIN
    ws.column_dimensions["B"].width = 8    # Symbol
    ws.column_dimensions["C"].width = 6    # KZ
    ws.column_dimensions["D"].width = 30   # Name
    ws.column_dimensions["E"].width = 14   # Ertragsverw.
    ws.column_dimensions["F"].width = 5    # Ccy
    ws.column_dimensions["G"].width = 10   # Shares
    ws.column_dimensions["H"].width = 13   # AE/share (native)
    ws.column_dimensions["I"].width = 13   # WA/share (native)
    ws.column_dimensions["J"].width = 13   # AE total (native)
    ws.column_dimensions["K"].width = 11   # FX rate
    ws.column_dimensions["L"].width = 14   # AE (EUR)
    ws.column_dimensions["M"].width = 14   # WA (EUR)
    ws.column_dimensions["N"].width = 14   # KeSt gross (EUR)
    ws.column_dimensions["O"].width = 14   # KeSt net (EUR)
    ws.column_dimensions["P"].width = 16   # AK-Korrektur (EUR)
    ws.column_dimensions["Q"].width = 12   # Meldedatum

    headers = [
        "ISIN", "Symbol", "KZ", "Name", "Ertragsverw.", "Ccy", "Shares",
        "AE/share", "WA/share",
        f"AE ({summary.tax_year} native)", "FX rate",
        "AE (EUR)", "WA (EUR)", "KeSt 27.5% gross", "KeSt net (after WA)",
        "AK-Korrektur (EUR)", "Meldedatum",
    ]

    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _hfill(SECTION_FILL)
        cell.alignment = _center()
        cell.border = _border()

    for r in summary.meldefonds:
        def fv(v):
            return float(v) if v is not None else None

        row_data = [
            r.isin,
            r.symbol,
            r.kz,
            r.name,
            r.ertragsverwendung,
            r.currency,
            float(r.shares_held),
            fv(r.ae_per_share),
            fv(r.wa_per_share),
            fv(r.ae_total_native),
            fv(r.fx_rate),
            fv(r.ae_total_eur),
            fv(r.wa_total_eur),
            fv(r.kest_gross_eur),
            fv(r.kest_net_eur),
            fv(r.ak_korrektur_eur),
            r.meldedatum or "",
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        ws.cell(row_idx, 2).font = _font(bold=True)   # symbol bold
        for col_idx in [8, 9, 10]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.0000'
        ws.cell(row_idx, 11).number_format = '0.0000'  # FX rate
        for col_idx in [12, 13, 14, 15, 16]:
            ws.cell(row_idx, col_idx).number_format = '#,##0.00 "EUR"'
        ws.cell(row_idx, 14).font = _font(bold=True)
        ws.cell(row_idx, 15).fill = _hfill(WARN_FILL)
        ws.cell(row_idx, 15).font = _font(bold=True)
        if r.ak_korrektur_eur and r.ak_korrektur_eur < 0:
            ws.cell(row_idx, 16).font = OFont(color="0070C0")  # blue for negative correction

        if r.warning:
            ws.cell(row_idx, 2).font = OFont(color="C00000", bold=True)

        for col_idx in range(1, 18):
            ws.cell(row_idx, col_idx).border = _border()

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(total_row, 11).value = "TOTAL"
    ws.cell(total_row, 11).font = _font(bold=True)
    for col_idx, val in [
        (12, summary.meldefonds_ae_eur),
        (13, summary.meldefonds_wa_eur),
        (14, summary.meldefonds_kest_gross_eur),
        (15, summary.meldefonds_kest_net_eur),
    ]:
        ws.cell(total_row, col_idx).value = float(val)
        ws.cell(total_row, col_idx).number_format = '#,##0.00 "EUR"'
        ws.cell(total_row, col_idx).font = _font(bold=True)
        ws.cell(total_row, col_idx).fill = _hfill(WARN_FILL if col_idx >= 14 else LIGHT_FILL)
        ws.cell(total_row, col_idx).border = _border()

    # Notes
    note_row = total_row + 2
    notes = [
        f"KZ 937 (Ausschüttungsgleiche Erträge Ausland) = {float(summary.meldefonds_ae_eur):,.2f} EUR  →  KeSt netto = {float(summary.meldefonds_kest_net_eur):,.2f} EUR (nach WA-Abzug)",
        "WA = Withhaltungsabzug: vom Fonds intern einbehaltene Quellensteuern — reduziert die KeSt-Schuld auf AE.",
        "AK-Korrektur = Anschaffungskosten-Korrektur (lt. OeKB): positiv erhöht, negativ senkt die steuerliche Anschaffungskosten.",
        "AE-Werte aus data/oekb_ae.yaml — Daten auf my.oekb.at verifizieren. PLACEHOLDER-Einträge vor Abgabe befüllen!",
        "Ausschüttend: Ausschüttungen separat in KZ 898; AE hier = nur thesaurierter Anteil (ggf. 0).",
        "Thesaurierend: GESAMTE Fondserträge als AE — auch ohne Barausschüttung jährlich steuerpflichtig.",
    ]
    for i, note in enumerate(notes):
        ws.merge_cells(f"A{note_row + i}:Q{note_row + i}")
        c = ws.cell(note_row + i, 1)
        c.value = note
        c.font = OFont(color="666666", italic=True, size=9)

    ws.freeze_panes = "A2"

    ws.freeze_panes = "A2"

