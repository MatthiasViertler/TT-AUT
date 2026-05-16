"""Household combined tax summary.

Reads per-person summary.json files and produces a combined Excel workbook
with side-by-side KeSt / dividend / portfolio figures.

Usage (via main.py):
    python main.py --household matthias,jessie --year 2025
"""

import json
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Styling (mirrors writer.py palette) ──────────────────────────────────────

_HEADER_FILL = "1F4E79"
_ACCENT_FILL  = "2E75B6"
_LIGHT_FILL   = "D6E4F0"
_TOTAL_FILL   = "C5E0B4"
_WARN_FILL    = "FFD700"
_WHITE        = "FFFFFF"
_BLACK        = "000000"


def _font(bold=False, color=_BLACK, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal="center", vertical="center")


def _right():
    return Alignment(horizontal="right", vertical="center")


# ── JSON helpers ──────────────────────────────────────────────────────────────

def _summary_path(person: str, year: int, users_dir: Path) -> Optional[Path]:
    """Return path to summary.json, checking new layout (output/{year}/) then legacy."""
    new = users_dir / person / "output" / str(year) / f"{person}_{year}_summary.json"
    if new.exists():
        return new
    legacy = users_dir / person / "output" / f"{person}_{year}_summary.json"
    if legacy.exists():
        return legacy
    return None


def _load_summary(person: str, year: int, users_dir: Path) -> Optional[dict]:
    path = _summary_path(person, year, users_dir)
    if path is None:
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _check_freshness(person: str, year: int, users_dir: Path) -> Optional[str]:
    path = _summary_path(person, year, users_dir)
    if path is None:
        return (f"  [warn]   No {year} summary for '{person}' — "
                f"run: python main.py --person {person} --year {year}")
    age_days = (datetime.now().timestamp() - path.stat().st_mtime) / 86400
    if age_days > 7:
        return (f"  [warn]   '{person}' summary is {age_days:.0f} days old — "
                f"consider re-running: python main.py --person {person} --year {year}")
    return None


def _d(data: dict, key: str) -> Decimal:
    try:
        raw = data.get(key)
        return Decimal(str(raw)) if raw is not None else Decimal("0")
    except Exception:
        return Decimal("0")


# ── Excel writer ──────────────────────────────────────────────────────────────

_ROWS: list[tuple[str, str, bool]] = [
    # (label, json_key, is_separator)
    ("Dividends & Interest (KZ 863)",   "kz_863",            False),
    ("Interest income component",        "interest_eur",      False),
    ("",                                 "",                  True),
    ("Domestic Gains (KZ 981)",          "kz_981",            False),
    ("Domestic Losses (KZ 891)",         "kz_891",            False),
    ("Foreign Gains (KZ 994)",           "kz_994",            False),
    ("Foreign Losses (KZ 892)",          "kz_892",            False),
    ("",                                 "",                  True),
    ("Net Taxable",                      "net_taxable_eur",   False),
    ("KeSt 27.5%",                       "kest_due_eur",      False),
    ("WHT Credited",                     "wht_creditable_eur",False),
    ("KeSt Remaining",                   "kest_remaining_eur",False),
    ("",                                 "",                  True),
    ("Portfolio Value (EUR)",            "portfolio_eur_computed", False),
    ("Dividend Yield (%)",               "dividend_yield_computed",False),
]

_LOSS_FIELDS = {"kz_891", "kz_892", "total_losses_eur"}
_HIGHLIGHT_FIELDS = {"kest_remaining_eur"}


def _write_excel(wb, persons: list[str], summaries: list[dict], year: int) -> None:
    ws = wb.active
    ws.title = "Household Summary"
    ws.column_dimensions["A"].width = 32
    for col in range(2, 2 + len(persons) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    row = [1]

    def _r():
        r = row[0]; row[0] += 1; return r

    # Title
    ncols = 1 + len(persons) + 1
    r = _r()
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    names = " + ".join(p.title() for p in persons)
    c = ws[f"A{r}"]
    c.value = f"Household Tax Summary  |  {names}  |  {year}"
    c.font = _font(bold=True, color=_WHITE, size=12)
    c.fill = _hfill(_HEADER_FILL)
    c.alignment = _center()
    ws.row_dimensions[r].height = 24

    # Header row
    r = _r()
    headers = [""] + [p.title() for p in persons] + ["Combined"]
    for col_idx, label in enumerate(headers, 1):
        c = ws.cell(r, col_idx, label)
        c.font = _font(bold=True, color=_WHITE, size=10)
        c.fill = _hfill(_ACCENT_FILL)
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[r].height = 20

    # Data rows
    for label, key, is_sep in _ROWS:
        r = _r()
        if is_sep:
            ws.row_dimensions[r].height = 6
            continue

        is_highlight = key in _HIGHLIGHT_FIELDS
        is_loss = key in _LOSS_FIELDS
        row_fill = _TOTAL_FILL if is_highlight else None

        # Label cell
        c = ws.cell(r, 1, label)
        c.font = _font(bold=is_highlight, size=10)
        c.fill = _hfill(row_fill or _WHITE)
        c.border = _border()

        # Per-person values
        total = Decimal("0")
        for col_idx, data in enumerate(summaries, 2):
            if key == "dividend_yield_computed":
                raw = data.get(key)
                val = float(raw) if raw is not None else None
                cell = ws.cell(r, col_idx, f"{val:.2f}%" if val is not None else "—")
                cell.alignment = _right()
            else:
                val = float(_d(data, key))
                cell = ws.cell(r, col_idx, -val if is_loss else val)
                cell.number_format = '#,##0.00'
                cell.alignment = _right()
                total += _d(data, key)
            cell.font = _font(bold=is_highlight, size=10)
            cell.fill = _hfill(row_fill or _WHITE)
            cell.border = _border()

        # Combined total cell
        col_total = 1 + len(persons) + 1
        if key == "dividend_yield_computed":
            # weighted average yield: sum(dividends) / sum(portfolio) * 100
            total_port = sum(_d(d, "portfolio_eur_computed") for d in summaries)
            total_div  = sum(_d(d, "total_dividends_eur") for d in summaries)
            avg_yield  = float(total_div / total_port * 100) if total_port else 0.0
            ct = ws.cell(r, col_total, f"{avg_yield:.2f}%")
            ct.alignment = _right()
        else:
            ct = ws.cell(r, col_total, float(-total if is_loss else total))
            ct.number_format = '#,##0.00'
            ct.alignment = _right()
        ct.font = _font(bold=is_highlight, size=10)
        ct.fill = _hfill(row_fill or _LIGHT_FILL)
        ct.border = _border()
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A3"


# ── Public entry point ────────────────────────────────────────────────────────

def run_household(persons: list[str], year: int, users_dir: Path) -> None:
    """Check freshness, load summaries, write combined Excel."""
    # Freshness check
    stale_warnings: list[str] = []
    for person in persons:
        warn = _check_freshness(person, year, users_dir)
        if warn:
            stale_warnings.append(warn)

    if stale_warnings:
        for w in stale_warnings:
            print(w, file=sys.stderr)
        # Only block if a summary is completely missing
        missing = [p for p in persons if _summary_path(p, year, users_dir) is None]
        if missing:
            print(f"  [error]  Cannot build household report — run per-person first.",
                  file=sys.stderr)
            sys.exit(1)

    # Load summaries
    summaries: list[dict] = []
    for person in persons:
        data = _load_summary(person, year, users_dir)
        if data is None:
            print(f"  [error]  Could not load summary for '{person}' {year}.",
                  file=sys.stderr)
            sys.exit(1)
        summaries.append(data)

    if not OPENPYXL_AVAILABLE:
        print("  [error]  openpyxl not installed — cannot write Excel.", file=sys.stderr)
        sys.exit(1)

    # Write Excel
    wb = openpyxl.Workbook()
    _write_excel(wb, persons, summaries, year)

    out_dir = users_dir / "household" / "output" / str(year)
    out_dir.mkdir(parents=True, exist_ok=True)
    slug = "_".join(persons)
    out_path = out_dir / f"household_{year}_{slug}.xlsx"
    wb.save(out_path)
    print(f"  [household] Written: {out_path}")
