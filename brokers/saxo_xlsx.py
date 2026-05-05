"""
SAXO Bank broker parser.
Handles two export types from the SAXO Reports section:

  AggregatedAmounts_{account}_{start}_{end}.xlsx  — trades + dividends
  ShareDividends_{account}_{start}_{end}.xlsx      — dividends with WHT detail

Dividends: prefer ShareDividends (has WHT %, quantity, original currency).
  AggregatedAmounts is the fallback if ShareDividends is not available.
  Do NOT pass both for the same period — dividends would double-count.

Trades: parsed from AggregatedAmounts 'Cash Movements' sheet.
  Sign of Amount Client Currency determines direction:
    positive = SELL (cash in), negative = BUY (cash out)
  No quantity data is available in SAXO exports. Each trade is stored as
  quantity=1, price=total_eur. FIFO matching therefore requires
  manual_cost_basis entries for any position not fully opened and closed
  within the provided files. Unmatched sells will trigger the standard warning.

Currency handling:
  2024+ DK accounts: Amount Client Currency is already EUR — used directly.
  2020 SG account:   Amount Client Currency is SGD — set as orig_currency so
                     the FX pipeline converts it to EUR via ECB rates.

Security: IBAN and sub-account IDs in these files are PII — never logged.
"""

import logging
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from core.models import (
    AssetClass, Domicile, NormalizedTransaction, TransactionType,
)

log = logging.getLogger(__name__)
ZERO = Decimal("0")

# SAXO exchange MIC suffix → ISO 3166-1 alpha-2 country code
_EXCHANGE_COUNTRY: dict[str, str] = {
    "xnas": "US",   # NASDAQ
    "xnys": "US",   # NYSE
    "xase": "US",   # NYSE American
    "xhkg": "CN",   # Hong Kong
    "xlon": "GB",   # London Stock Exchange
    "xcse": "DK",   # Copenhagen
    "xsto": "SE",   # Stockholm
    "xfra": "DE",   # Frankfurt
    "xpar": "FR",   # Paris
    "xams": "NL",   # Amsterdam / Euronext
    "xswx": "CH",   # Swiss Exchange
    "xasx": "AU",   # ASX
    "xtse": "CA",   # Toronto
    "xjpx": "JP",   # Japan
    "xkrx": "KR",   # Korea
    "xtai": "TW",   # Taiwan
}

_SUBTYPE_ASSET_CLASS: dict[str, AssetClass] = {
    "Stock":              AssetClass.STOCK,
    "Etf":                AssetClass.ETF,
    "ExchangeTradedFund": AssetClass.ETF,
    "Bond":               AssetClass.BOND,
}

# Amount Type Name values that carry no Austrian tax consequence
_SKIP_TYPES = {
    "Custody Fee", "Client Custody Fee", "VAT", "GST on Commission",
    "Cash Amount",                              # deposits / withdrawals
    "ADR Charges", "Depository Charges",
    "Net P/L",                                  # unrealized daily mark
    "Position Exposure", "Percent return per Instrument",
    "Position Values", "Change in Accruals", "Accruals", "Cash",
    "Corporate Actions - Fractions",            # tiny fractional-share cash
    "Corporate Actions - Stock Dividend",       # stock dividend (no cash)
}

_CURRENCY_RE = re.compile(r'\b([A-Z]{3})\b')
_AMOUNT_RE   = re.compile(r'-?\s*\d[\d,\.]*')


# ── Public API ────────────────────────────────────────────────────────────────

def detect(path: Path) -> bool:
    """Return True if this file is a SAXO xlsx export."""
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and (name.startswith("aggregatedamounts_") or name.startswith("sharedividends_"))
    )


def get_account_id(path: Path) -> Optional[str]:
    """Extract account ID from filename pattern {Type}_{AccountID}_{...}.xlsx."""
    parts = path.stem.split("_")
    return parts[1] if len(parts) >= 2 else None


def parse(path: Path, config: dict) -> tuple[list[NormalizedTransaction], Optional[str]]:
    """Parse a SAXO xlsx export. Returns (transactions, account_id)."""
    account_id = get_account_id(path)
    name = path.name.lower()

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        log.error("openpyxl not installed — cannot parse SAXO xlsx. "
                  "Run: pip install openpyxl")
        return [], account_id

    wb = load_workbook(path, data_only=True)
    try:
        if name.startswith("sharedividends_"):
            txns = _parse_share_dividends(wb, path, account_id)
        elif name.startswith("aggregatedamounts_"):
            txns = _parse_aggregated_amounts(wb, path, account_id)
        else:
            log.warning(f"SAXO parser: unrecognised filename pattern: {path.name}")
            txns = []
    finally:
        wb.close()

    log.info(f"SAXO parser: {path.name} → {len(txns)} transactions")
    return txns, account_id


# ── AggregatedAmounts parser ──────────────────────────────────────────────────

def _parse_aggregated_amounts(wb, path: Path,
                               account_id: Optional[str]) -> list[NormalizedTransaction]:
    if "Cash Movements" not in wb.sheetnames:
        log.warning(f"SAXO: 'Cash Movements' sheet not found in {path.name}")
        return []

    ws = wb["Cash Movements"]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return []

    headers = [str(h) if h is not None else "" for h in all_rows[0]]

    # Detect account base currency from first data row
    first = dict(zip(headers, all_rows[1]))
    client_ccy = first.get("Client Currency") or "EUR"
    is_eur_account = (client_ccy == "EUR")

    # Group rows by (date, UIC) to associate commissions with trades
    # and WHT with dividends.
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for raw in all_rows[1:]:
        d = dict(zip(headers, raw))
        amt_type = str(d.get("Amount Type Name") or "")
        if amt_type in _SKIP_TYPES:
            continue
        if amt_type not in {
            "Share Amount", "Commission", "Exchange Fee",
            "Corporate Actions - Cash Dividends",
            "Corporate Actions - Withholding Tax",
            "Corporate Actions - Fee",
        }:
            continue
        dt = _parse_date(d.get("Date"))
        if dt is None:
            continue
        uic = d.get("Unified Instrument Code (UIC)") or 0
        groups[(dt, uic)].append(d)

    result: list[NormalizedTransaction] = []
    counters: dict[str, int] = defaultdict(int)

    for (dt, uic), group_rows in sorted(groups.items()):
        trade_rows = [r for r in group_rows
                      if r.get("Amount Type Name") == "Share Amount"]
        cost_rows  = [r for r in group_rows
                      if r.get("Amount Type Name") in
                      {"Commission", "Exchange Fee"}]
        div_rows   = [r for r in group_rows
                      if r.get("Amount Type Name") ==
                      "Corporate Actions - Cash Dividends"]
        wht_rows   = [r for r in group_rows
                      if r.get("Amount Type Name") ==
                      "Corporate Actions - Withholding Tax"]

        # ── Trades ────────────────────────────────────────────────────────────
        for row in trade_rows:
            sym, exchange = _parse_symbol(
                row.get("Instrument Symbol") or
                row.get("Underlying Instrument Symbol")
            )
            if sym == "UNKNOWN":
                # 2020 SG format: symbol is None, use description
                desc_raw = (row.get("Instrument Description") or
                            row.get("Underlying Instrument Description") or "")
                sym = _symbol_from_description(str(desc_raw), uic)
            description = (row.get("Instrument Description") or
                           row.get("Underlying Instrument Description") or sym)

            # Amount Client Currency: EUR for DK accounts, SGD for SG account
            amt = _d(row.get("Amount Client Currency"))
            # Total commission for this trade (same date+UIC)
            comm = sum((_d(r.get("Amount Client Currency")) for r in cost_rows), ZERO)

            orig_ccy = "EUR" if is_eur_account else client_ccy
            txn_type = TransactionType.SELL if amt > ZERO else TransactionType.BUY

            subtype = (row.get("Instrument SubType") or
                       row.get("Underlying Instrument SubType") or "")
            asset_cls = _SUBTYPE_ASSET_CLASS.get(str(subtype), AssetClass.STOCK)

            country = _EXCHANGE_COUNTRY.get(exchange, "US")

            base_id = f"saxo_{account_id}_{dt}_{uic}_trade"
            counters[base_id] += 1
            raw_id = base_id if counters[base_id] == 1 else f"{base_id}_{counters[base_id]}"

            result.append(NormalizedTransaction(
                broker="saxo",
                raw_id=raw_id,
                trade_date=dt,
                settle_date=None,
                txn_type=txn_type,
                asset_class=asset_cls,
                symbol=sym,
                isin=None,
                description=str(description)[:80],
                country_code=country,
                domicile=Domicile.FOREIGN,
                # No quantity in SAXO exports — store as 1 lot at total price
                quantity=Decimal("1"),
                price=amt.copy_abs(),
                price_currency=orig_ccy,
                orig_currency=orig_ccy,
                orig_amount=amt,
                wht_amount_orig=ZERO,
                commission=comm.copy_abs(),
                commission_currency=orig_ccy,
                source_file=path.name,
            ))

        # ── Dividends (fallback — prefer ShareDividends file) ─────────────────
        if not div_rows:
            continue

        for row in div_rows:
            sym, exchange = _parse_symbol(
                row.get("Instrument Symbol") or
                row.get("Underlying Instrument Symbol")
            )
            if sym == "UNKNOWN":
                desc_raw = (row.get("Instrument Description") or
                            row.get("Underlying Instrument Description") or "")
                sym = _symbol_from_description(str(desc_raw), uic)
            description = (row.get("Instrument Description") or
                           row.get("Underlying Instrument Description") or sym)

            gross = _d(row.get("Amount Client Currency"))
            wht   = sum((_d(r.get("Amount Client Currency")).copy_abs()
                         for r in wht_rows), ZERO)

            orig_ccy = "EUR" if is_eur_account else client_ccy
            country  = _EXCHANGE_COUNTRY.get(exchange, "US")

            base_id = f"saxo_{account_id}_{dt}_{uic}_div"
            counters[base_id] += 1
            raw_id = base_id if counters[base_id] == 1 else f"{base_id}_{counters[base_id]}"

            result.append(NormalizedTransaction(
                broker="saxo",
                raw_id=raw_id,
                trade_date=dt,
                settle_date=None,
                txn_type=TransactionType.DIVIDEND,
                asset_class=asset_cls if not div_rows else AssetClass.STOCK,
                symbol=sym,
                isin=None,
                description=str(description)[:80],
                country_code=country,
                domicile=Domicile.FOREIGN,
                quantity=None,
                price=None,
                price_currency=None,
                orig_currency=orig_ccy,
                orig_amount=gross,
                wht_amount_orig=wht,
                source_file=path.name,
            ))

    return result


# ── ShareDividends parser ─────────────────────────────────────────────────────

def _parse_share_dividends(wb, path: Path,
                            account_id: Optional[str]) -> list[NormalizedTransaction]:
    if "Share Dividends" not in wb.sheetnames:
        log.warning(f"SAXO: 'Share Dividends' sheet not found in {path.name}")
        return []

    ws = wb["Share Dividends"]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return []

    headers = [str(h) if h is not None else "" for h in all_rows[0]]

    # Group by (symbol, pay_date, holding) to merge split rows.
    # SAXO sometimes splits one dividend event into:
    #   row A: gross=0, WHT=full_amount
    #   row B: gross=full_amount, WHT=0
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for raw in all_rows[1:]:
        d = dict(zip(headers, raw))
        sym_raw = str(d.get("Instrument Symbol") or "").strip()
        pay_dt  = _parse_date(d.get("Pay Date") or d.get("Posting Date"))
        holding = d.get("Holding") or 0
        if sym_raw and pay_dt:
            groups[(sym_raw, pay_dt, holding)].append(d)

    result: list[NormalizedTransaction] = []
    counters: dict[str, int] = defaultdict(int)

    for (sym_raw, pay_dt, _holding), group in sorted(groups.items(),
                                                       key=lambda kv: (kv[0][1], kv[0][0])):
        sym, exchange = _parse_symbol(sym_raw)

        gross_total = ZERO
        wht_total   = ZERO
        fee_total   = ZERO
        orig_ccy    = "USD"
        description = sym

        for d in group:
            description = str(d.get("Instrument") or sym)
            div_str = str(d.get("Dividend amount") or "")
            wht_str = str(d.get("Withholding tax amount") or "")
            fee_str = str(d.get("Fee amount") or "")

            gross_total += _parse_ccy_string(div_str)
            wht_total   += _parse_ccy_string(wht_str).copy_abs()
            fee_total   += _parse_ccy_string(fee_str).copy_abs()

            ccy = _extract_ccy(div_str) or _extract_ccy(wht_str)
            if ccy:
                orig_ccy = ccy

        if gross_total == ZERO and wht_total == ZERO:
            continue

        country = _EXCHANGE_COUNTRY.get(exchange, "US")

        base_id = f"saxo_{account_id}_{pay_dt}_{sym}_div"
        counters[base_id] += 1
        raw_id = base_id if counters[base_id] == 1 else f"{base_id}_{counters[base_id]}"

        result.append(NormalizedTransaction(
            broker="saxo",
            raw_id=raw_id,
            trade_date=pay_dt,
            settle_date=None,
            txn_type=TransactionType.DIVIDEND,
            asset_class=AssetClass.STOCK,
            symbol=sym,
            isin=None,
            description=description[:80],
            country_code=country,
            domicile=Domicile.FOREIGN,
            quantity=None,
            price=None,
            price_currency=None,
            orig_currency=orig_ccy,
            orig_amount=gross_total,
            wht_amount_orig=wht_total,
            commission=fee_total,          # ADR/depository fees
            commission_currency=orig_ccy,
            source_file=path.name,
        ))

    return result


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(v) -> Optional[date]:
    """Parse datetime object, DD-MM-YYYY string, or YYYY-MM-DD string → date."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # DD-MM-YYYY (SAXO 2024+ format)
    if len(s) == 10 and s[2] == "-" and s[5] == "-":
        try:
            return date(int(s[6:10]), int(s[3:5]), int(s[0:2]))
        except ValueError:
            pass
    # YYYY-MM-DD fallback
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _parse_symbol(raw) -> tuple[str, str]:
    """'NVDA:xnas' → ('NVDA', 'xnas').  Returns ('UNKNOWN', '') on failure."""
    if not raw:
        return "UNKNOWN", ""
    s = str(raw).strip()
    if ":" in s:
        sym, exc = s.rsplit(":", 1)
        return sym.strip().upper(), exc.strip().lower()
    return s.upper(), ""


def _symbol_from_description(description: str, uic) -> str:
    """Fallback symbol for 2020 SG format where Instrument Symbol is absent.
    Uses UIC-based name since description words ('Advanced', 'The', …) are not
    reliable tickers. Add symbol_aliases in config.local.yaml to remap if needed.
    """
    return f"UIC{uic}"


def _d(v) -> Decimal:
    if v is None:
        return ZERO
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return ZERO


def _parse_ccy_string(s: str) -> Decimal:
    """'USD 23.00' | '-USD 3.45' | 'USD -3.45' → Decimal. Returns ZERO on failure."""
    if not s or s.strip() in ("", "None"):
        return ZERO
    # Strip currency code letters, keep sign and digits
    numeric = re.sub(r"[A-Za-z\s]", "", s).replace(",", "")
    try:
        return Decimal(numeric)
    except InvalidOperation:
        return ZERO


def _extract_ccy(s: str) -> Optional[str]:
    """Extract first 3-letter currency code from 'USD 23.00' → 'USD'."""
    m = _CURRENCY_RE.search(s)
    return m.group(1) if m else None
