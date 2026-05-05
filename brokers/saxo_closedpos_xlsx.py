"""
SAXO Bank ClosedPositions xlsx parser.

Handles: ClosedPositions_{account}_{start}_{end}.xlsx exports from the
SAXO Reports section (Account → Report → Closed Positions, Excel format).

Capital gains (BUY + SELL):
  Each row is one close event. The parser emits:
  - SELL with real QuantityClose (abs value)
  - BUY  with real Quantity Open, deduped by OpenPositionId

  Deduplication: raw_id = "saxo_cp_buy_{OpenPositionId}" is globally unique
  per lot. The pipeline's existing raw_id dedup handles the case where the
  same lot appears in multiple ClosedPositions files (partial closes across
  years) — only the first occurrence is kept.

SG-transferred positions (skip list):
  SAXO carried over SG-account lots to the DK account at the transfer date
  (2024-03-07 for Matthias). The ClosedPositions Open Price is the SG avg
  cost, but Trade Date Open is the transfer date — not the original purchase
  date. Using that date for ECB FX conversion gives wrong EUR cost basis for
  Austrian KeSt.

  Fix: list these open dates in saxo_closedpos_skip_buy_open_dates.  The
  parser then emits SELL with qty=1 (compatible with manual_cost_basis qty=1
  convention) and no BUY (manual_cost_basis handles the cost side).

  NVDA note: the 10:1 split on 2024-06-10 created a synthetic lot in SAXO's
  system at the split-adjusted price. Add "2024-06-10" to the skip list so
  the existing NVDA manual_cost_basis entries (qty=1) handle the cost basis.

Config keys (top-level, in config.local.yaml):
  saxo_closedpos_skip_buy_open_dates: ["2024-03-07", "2024-06-10"]
    ISO date strings. Any row whose Trade Date Open matches is treated as
    an SG-transferred or synthetic lot: SELL qty=1, BUY suppressed.

  saxo_skip_agg_trades: true
    Set this in config.local.yaml alongside ClosedPositions use to prevent
    AggregatedAmounts from also emitting BUY/SELL rows for the same positions.
    AggregatedAmounts is then used for dividends only.

Commissions:
  ClosedPositions carries no per-trade commission breakdown. Commission is
  set to ZERO. For most SAXO trades this has negligible tax impact.

Security: account IDs in filenames and 'Account ID' column are PII — never log.
"""

import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from core.models import AssetClass, Domicile, NormalizedTransaction, TransactionType
from brokers.saxo_xlsx import _EXCHANGE_COUNTRY, _parse_symbol

log = logging.getLogger(__name__)
ZERO = Decimal("0")

_SKIP_ASSET_TYPES = {"StockOption", "ContractFutures", "FxSpot"}

_ASSET_CLASS_MAP: dict[str, AssetClass] = {
    "Stock": AssetClass.STOCK,
    "Etf":   AssetClass.ETF,
    "Bond":  AssetClass.BOND,
}


# ── Public API ────────────────────────────────────────────────────────────────

def detect(path: Path) -> bool:
    """Return True if this file is a SAXO ClosedPositions xlsx export."""
    name = path.name.lower()
    return path.suffix.lower() == ".xlsx" and name.startswith("closedpositions_")


def get_account_id(path: Path) -> Optional[str]:
    """Extract account ID from filename ClosedPositions_{AccountID}_{...}.xlsx."""
    parts = path.stem.split("_")
    return parts[1] if len(parts) >= 2 else None


def parse(path: Path, config: dict) -> tuple[list[NormalizedTransaction], Optional[str]]:
    """Parse a SAXO ClosedPositions xlsx. Returns (transactions, account_id)."""
    account_id = get_account_id(path)

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        log.error("openpyxl not installed — cannot parse SAXO xlsx. Run: pip install openpyxl")
        return [], account_id

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        return [], account_id

    headers = [str(h) if h is not None else "" for h in rows[0]]

    # Dates whose open leg must come from manual_cost_basis (SG transfer / synthetic lots)
    skip_open_dates: set[date] = _load_skip_dates(config)

    result: list[NormalizedTransaction] = []
    seen_open_pos_ids: set = set()  # within-file dedup; cross-file handled by pipeline raw_id

    for raw in rows[1:]:
        d = dict(zip(headers, raw))

        asset_type = str(d.get("Asset type") or "")
        if asset_type in _SKIP_ASSET_TYPES:
            continue

        close_dt = _parse_date(d.get("Trade Date Close"))
        open_dt  = _parse_date(d.get("Trade Date Open"))
        if close_dt is None or open_dt is None:
            continue

        sym, exchange = _parse_symbol(str(d.get("Instrument Symbol") or ""))
        if sym == "UNKNOWN":
            log.warning(f"ClosedPositions {path.name}: no symbol in row, skipping")
            continue

        instr_ccy   = str(d.get("Instrument currency") or "USD")
        qty_close   = _d(d.get("QuantityClose"))   # negative (short side)
        qty_open    = _d(d.get("Quantity Open"))   # positive
        open_px     = _d(d.get("Open Price"))
        close_px    = _d(d.get("Close Price"))
        open_pos_id = d.get("OpenPositionId")
        close_pos_id= d.get("ClosePositionId")

        sell_qty = qty_close.copy_abs()
        if sell_qty == ZERO:
            continue

        country   = _EXCHANGE_COUNTRY.get(exchange, "US")
        asset_cls = _ASSET_CLASS_MAP.get(asset_type, AssetClass.STOCK)
        desc      = str(d.get("Instrument Description") or sym)[:80]

        is_sg_lot = open_dt in skip_open_dates

        # ── SELL ──────────────────────────────────────────────────────────────
        if is_sg_lot:
            # Emit qty=1 so the existing manual_cost_basis (also qty=1) is consumed.
            # orig_amount = total proceeds for this close event in instrument currency.
            emit_sell_qty = Decimal("1")
        else:
            emit_sell_qty = sell_qty

        result.append(NormalizedTransaction(
            broker="saxo_cp",
            raw_id=f"saxo_cp_sell_{close_pos_id}",
            trade_date=close_dt,
            settle_date=None,
            txn_type=TransactionType.SELL,
            asset_class=asset_cls,
            symbol=sym,
            isin=None,
            description=desc,
            country_code=country,
            domicile=Domicile.FOREIGN,
            quantity=emit_sell_qty,
            price=close_px,
            price_currency=instr_ccy,
            orig_currency=instr_ccy,
            orig_amount=close_px * sell_qty,   # total proceeds, always real amount
            wht_amount_orig=ZERO,
            commission=ZERO,
            commission_currency=instr_ccy,
            source_file=path.name,
        ))

        # ── BUY (DK-native lots only) ─────────────────────────────────────────
        # Suppressed for SG-transferred lots; manual_cost_basis handles cost side.
        # Deduped within file (seen_open_pos_ids) and across files (pipeline raw_id).
        if not is_sg_lot and open_pos_id not in seen_open_pos_ids:
            seen_open_pos_ids.add(open_pos_id)
            result.append(NormalizedTransaction(
                broker="saxo_cp",
                raw_id=f"saxo_cp_buy_{open_pos_id}",
                trade_date=open_dt,
                settle_date=None,
                txn_type=TransactionType.BUY,
                asset_class=asset_cls,
                symbol=sym,
                isin=None,
                description=desc,
                country_code=country,
                domicile=Domicile.FOREIGN,
                quantity=qty_open,
                price=open_px,
                price_currency=instr_ccy,
                orig_currency=instr_ccy,
                orig_amount=-(qty_open * open_px),  # negative = cash out
                wht_amount_orig=ZERO,
                commission=ZERO,
                commission_currency=instr_ccy,
                source_file=path.name,
            ))

    log.info(f"ClosedPositions parser: {path.name} → {len(result)} transactions")
    return result, account_id


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_skip_dates(config: dict) -> set[date]:
    raw = config.get("saxo_closedpos_skip_buy_open_dates", [])
    result: set[date] = set()
    for d in raw:
        if isinstance(d, date) and not isinstance(d, datetime):
            result.add(d)
        elif isinstance(d, datetime):
            result.add(d.date())
        else:
            try:
                result.add(date.fromisoformat(str(d)))
            except ValueError:
                log.warning(f"ClosedPositions: invalid skip date '{d}' in config")
    return result


def _parse_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if len(s) == 10 and s[2] == "-" and s[5] == "-":
        try:
            return date(int(s[6:10]), int(s[3:5]), int(s[0:2]))
        except ValueError:
            pass
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _d(v) -> Decimal:
    if v is None:
        return ZERO
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return ZERO
